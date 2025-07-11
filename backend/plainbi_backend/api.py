# -*- coding: utf-8 -*-
"""
Created on Thu Mar  9 11:19:08 2023

@author: kribbel

how to run
first window
~/plainbi/frontend> npm start
second window
~/plainbi/backend> python plainbi_backend.py
3.4.2025

in Browser:
http://localhost:3001/
# swagger testing
http://localhost:3001/apidocs/

"""

import os
import sys
import logging
import traceback
import tempfile
from datetime import date,datetime

import base64
import hashlib
import requests
#import urllib
from urllib.parse import urlparse, parse_qs
import json

import pprint
#import sqlalchemy
from sqlalchemy.exc import SQLAlchemyError
import decimal
import csv
import pandas as pd
from flask_bcrypt import Bcrypt
from openpyxl import load_workbook
#from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table #, TableStyleInfo
import smtplib
import pandas.io.formats.excel as fmt_xl
import ast

from dotenv import load_dotenv

from functools import wraps
from flask import Flask, jsonify, request, Response, Blueprint, make_response, session, url_for
from flask_session import Session
#from flask_cors import CORS

#from flask.json import JSONEncoder
from json import JSONEncoder
import jwt
from jwt import PyJWKClient
import secrets

with_swagger = False
if "PLAINBI_NOSWAGGER" in list(dict(os.environ).keys()):
    with_swagger = False
else:
    try:
        from flasgger import Swagger
        with_swagger = True
    except Exception as e_swagger:
        with_swagger = False

from plainbi_backend.utils import db_subs_env, prep_pk_from_url, is_id, last_stmt_has_errors, make_pk_where_clause, urlsafe_decode_params, pre_jsonify_items_transformer, parse_filter, dbg, err, warn, dbg_api_call
from plainbi_backend.db import sql_select, get_item_raw, get_metadata_raw, db_connect, db_connect_test, db_exec, db_ins, db_upd, db_del, get_current_timestamp, get_next_seq, repo_lookup_select, get_repo_adhoc_sql_stmt, get_repo_customsql_sql_stmt, get_profile, add_auth_to_where_clause, add_offset_limit, audit, db_adduser, db_passwd, get_db_type, get_dbversion, load_datasources_from_repo, get_db_by_id_or_alias
from plainbi_backend.repo import create_repo_db

# import the global variable config
import plainbi_backend.config as cfg
from plainbi_backend.config import config


#log = logging.getLogger(config.logger_name)
log = logging.getLogger(__name__)

try:
    import ldap3
    config.with_ldap3=True
    log.info("LDAP3 enabled")
except:
    print("LDAP disabled because not installed")
    config.with_ldap3=False
    log.info("LDAP disabled because not installed")

# try to load identiy for microsoft sso authentication if available
auth = None
try:
    #from msal import ConfidentialClientApplication, ClientCredential
    import msal
    config.with_sso=True
    print("Microsoft SSO enabled")
    log.info("Microsoft SSO enabled")
except:
    config.with_sso=False
    print("Microsoft SSO disabled because not installed")
    log.info("Microsoft SSO disabled because not installed")

api = Blueprint('api', __name__)

def myjsonify(d: dict):
    try:
        jd=jsonify(d)
    except Exception as e:
        err("cannot jsonify dict %s",str(d))
        jd=jsonify({"error":"cannot jsonify output", "detail":str(e)})
        log.exception(e)
        dbg("set status_code to 500 due to jsonify error")
        #jd.status_code=500
    if config.dbg_level >= 3 and log.getEffectiveLevel() == logging.DEBUG:
        dbg("--- myjsonify json output")
        pprint.pprint(jd)
        pprint.pprint(d)
        dbg("--- end myjsonify json output")
    return jd

class CustomJSONEncoder(JSONEncoder):
    def default(self, obj):
        dbg("CustomJSONEncoder %s / %s",str(type(obj)),str(obj))
        try:
            if isinstance(obj, datetime):
                return obj.strftime("%Y-%m-%d %H:%M:%S.%f")
            elif isinstance(obj, date):
                return obj.strftime("%Y-%m-%d")
            elif isinstance(obj,decimal.Decimal):
                return str(obj)
            elif isinstance(obj, Exception):
                return str(obj)
            iterable = iter(obj)
        except TypeError:
            pass
        else:
            return list(iterable)
        return JSONEncoder.default(self, obj)

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        dbg("token req")
        token = request.headers.get('Authorization')
        dbg("token=%s",str(token))

        if not token:
            return myjsonify({'message': 'Token is missing'}), 401

        try:
            tokdata = jwt.decode(token, config.SECRET_KEY, algorithms=['HS256'])
            dbg("data2=%s",str(tokdata))
            #config.current_user=tokdata['username']
            #dbg("cur user=%s",str(config.current_user))
        except jwt.ExpiredSignatureError:
            return myjsonify({'message': 'Token has expired'}), 401
        except jwt.InvalidTokenError:
            return myjsonify({'message': 'Invalid token x'}), 401
        return f(tokdata, *args, **kwargs)
    return decorated


repo_table_prefix="plainbi_"

api_root="/api"
api_prefix=api_root+"/crud"
repo_api_prefix=api_root+"/repo"
api_metadata_prefix=api_root+"/metadata"
cursor_desc_fields=["name","type_code","display_size","internal_size","precision","scale","null_ok"]

nodb_msg = { "error" :"-no-open-db", "message":"Datenbankverbindung ungültig in datasources, prüfe die Repo Konfiguration" }

metadata_tab_query="""
SELECT 
    DB_NAME() AS database_name,
    SCHEMA_NAME(t.schema_id) AS schema_name,
    t.name AS table_name,
    DB_NAME()+'.'+SCHEMA_NAME(t.schema_id)+'.'+t.name AS full_table_name
FROM sys.tables t
ORDER BY database_name, schema_name, table_name
"""

#
@api.route('/', methods=['GET'])
def welcome():
    """
    welcome message to the backend rest server if no specific url is given

    ---
    tags:
      - Misc
    produces:
      - text/html
    responses:
      200:
        description: Successful operation
        examples:
           text/html:
                <html>
                <body>
                <h1>Welcome to PLAINBI Backend</h1>
                <p>Version 0.7 29.07.2024</p>
                <p>Repo Database version PostgreSQL 15.7 on x86_64-pc-linux-gnu, compiled by gcc (GCC) 8.5.0 20210514 (Red Hat 8.5.0-22), 64-bit</p>
                <p>If you want to initialize the repository click <a href="/api/repo/init_repo">here</a></p>
                </body>
                </html>
    """
    dbversion=get_dbversion(config.repoengine)
    p=f"""
    <html>
    <body>
    <h1>Welcome to PLAINBI Backend</h1>
    <p>Version {config.version}</p>
    <p>Repo Database version {dbversion}</p>
    <p>If you want to initialize the repository click <a href="{repo_api_prefix+'/init_repo'}">here</a></p>
    </body>    
    </html>    
    """
    return p

@api.route('/version', methods=['GET'])
@api.route(api_root+'/version', methods=['GET'])
def get_version():
    """
    return the version number of the backend
    ---
    tags:
      - Misc
    produces:
      - text/plain
    responses:
      200:
        description: Successful operation
        examples:
          text/plain: '0.7 vom 5.8.2024'
    """
    return config.version

@api.route(api_root+'/backend_version', methods=['GET'])
@api.route(api_root+'/db_version', methods=['GET'])
@api.route(api_root+'/dbversion', methods=['GET'])
def get_backend_version():
    """
    return the database type and version of the backend

    ---
    tags:
      - Misc
    produces:
      - text/plain
    responses:
      200:
        description: Successful operation
        examples:
          text/plain: '
            Plainbi Backend: 0.7 29.07.2024 
            Repository: PostgreSQL 15.7 on x86_64-pc-linux-gnu, compiled by gcc (GCC) 8.5.0 20210514 (Red Hat 8.5.0-22), 64-bit'
    """
    dbg_api_call(request)
    dbversion=get_dbversion(config.repoengine)
    return "Plainbi Backend: "+config.version+"\nRepository: "+str(dbversion)

@api.route(api_root+'/loglevel/<loglevel>', methods=['GET'])
def set_log_level(loglevel):
    """
    set log level log.setLevel(
    """
    if len(request.args) > 0:
        for key, value in request.args.items():
            log.info("loglevel arg: %s val: %s",key,value)
            if key=="loggers":
                lognames=value.split(",")
                dbg("loggers are: "+str(lognames))
                loggers = [logging.getLogger(name) for name in lognames]
    else:
        loggers = [logging.getLogger(name) for name in logging.root.manager.loggerDict if "plainbi" in name]
        dbg("all plainbi loggers")

    for l in loggers:
      if loglevel=="INFO":
        l.setLevel(logging.INFO)
        log.info(f"LogLevel {loglevel} for {l.name} enabled")
        config.dbg_level = 1
      if loglevel=="DEBUG":
        l.setLevel(logging.DEBUG)
        log.info(f"LogLevel {loglevel} for {l.name} enabled")
        config.dbg_level = 1
      if loglevel=="DEBUG1":
        l.setLevel(logging.DEBUG)
        log.info(f"LogLevel {loglevel} for {l.name} enabled")
        config.dbg_level = 1
      if loglevel=="DEBUG2":
        l.setLevel(logging.DEBUG)
        log.info(f"LogLevel {loglevel} for {l.name} enabled")
        config.dbg_level = 2
      if loglevel=="DEBUG3":
        l.setLevel(logging.DEBUG)
        log.info(f"LogLevel {loglevel} for {l.name} enabled")
        config.dbg_level = 3
    return 'set log level '+loglevel, 200

@api.route('/status', methods=['GET'])
@api.route(api_root+'/status', methods=['GET'])
def get_api_status():
    """
    return status of the backend
    ---
    tags:
      - Misc
    produces:
      - text/plain
    responses:
      200:
        description: Successful operation
        examples:
          text/plain: '0.7 vom 5.8.2024'
    """
    s=""
    s+="API Version: "+config.version+"\n"
    dbversion=get_dbversion(config.repoengine)+"\n"
    s+="Repository: "+str(dbversion)+"\n"

    for l in logging.root.manager.loggerDict:
        if "plainbi" in l:
            lg=logging.getLogger(l)
            s+=l+": "+logging.getLevelName(lg.getEffectiveLevel())
    
    s+="\nLog Level: "+str(config.dbg_level)+"\n"

    s+="\nall loggers: "+", ".join(logging.root.manager.loggerDict)

    return s


@api.route(api_root+'/email', methods=['POST'])
@token_required
def sndemail(tokdata):
    """
    send an smtp email

    needs environment variables SMTP_SERVER, SMTP_PORT, SMTP_USER, SMTP_PASSWORD

    ---
    tags:
      - Utils
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      -  name: body
         in: body
         description: all elements in body will be interpreted
         schema:
            required:
              - to
              - subject
              - body
            properties:
              to:
                type: string
                description: email to adress
                example: " "
              subject:
                type: string
                description: subject of email
              body:
                type: string
                description: email text
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Email wurde versendet
      500:
        description: sendmail error
        examples:
          application/json: 
            error: sendemail
            message: Email konnte nicht versendet werden
    """
    dbg("++++++++++ entering sndemail")
    audit(tokdata,request)
    out={}

    dbg("sndemail: parse request data")
    data_bytes = request.get_data()
    data_string = data_bytes.decode('utf-8')
    item = json.loads(data_string.strip("'"))
    
    dbg("sndemail: get smtp config")
    try:
        # SMTP server configuration
        smtp_server = os.environ["SMTP_SERVER"] # "smtp.gmail.com"
        smtp_port = int(os.environ["SMTP_PORT"]) 
        smtp_user = os.environ["SMTP_USER"] # "your_email@gmail.com"
        smtp_password = os.environ.get("SMTP_PASSWORD") # "your_password" or none if env does not exist
        if isinstance(smtp_password,str):
          if len(smtp_password)==0:
              smtp_password=None
    except Exception as e:
        err("sendmail error: %s", str(e))
        out["error"]="sendemail"
        out["message"]="Email Konfiguration invalid"
        return myjsonify(out), 500
        # Create the email headers and body

    dbg("sndemail: check email params")
    if "to" not in item.keys() or "subject" not in item.keys() or "body" not in item.keys():
        err("sendmail error: to, subject or body in request post arguments missing")
        out["error"]="sendemail"
        out["message"]="Email invalid"
        return myjsonify(out), 500
    email_message = f"From: {smtp_user}\nTo: {item['to']}\nSubject: {item['subject']}\n\n{item['body']}"
    dbg("sndemail: send email")
    try:
        # Connect to the SMTP server
        dbg("sndemail: connect to smtp server")
        server = smtplib.SMTP(smtp_server, smtp_port)
        # Log in to the server
        if smtp_password is not None:
            # login to mailserver if password is specified
            dbg("sndemail: login to smtp server (there is a password)")
            server.login(smtp_user, smtp_password)
        # Send the email
        dbg("sndemail: send the mail")
        server.sendmail(smtp_user, item["to"], email_message)
        # Disconnect from the server
        dbg("sndemail: quit from server")
        server.quit()
    except Exception as e:
        err("sendmail error: %s", str(e))
        out["error"]="sendemail"
        out["message"]="Email konnte nicht versendet werden"
        return myjsonify(out), 500
    out["message"]="Email wurde versendet"
    return myjsonify(out)


@api.route(api_root+'/distinctvalues/<db>/<tabnam>/<colnam>', methods=['GET'])
@token_required
def distinctvalues(tokdata,db,tabnam,colnam):
    """
    get distinct values of a column in a table

    returns json with keys "data", "total_count"

    ---
    tags:
      - CRUD
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: db
        in: path
        type: string
        required: true
        description: id or alias of the database connection defined in repository table plainbi_datasource (0=repository)
      - name: tabnam
        in: path
        type: string
        required: true
        description: name of table in database 
      - name: colnam
        in: path
        type: string
        required: true
        description: name of a column in the table 
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering distinctvalues")
    dbg("distinctvalues param tab is <%s>",str(tabnam))
    dbg("distinctvalues param col is <%s>",str(colnam))
    audit(tokdata,request)
    dbengine=get_db_by_id_or_alias(db)
    if dbengine is None:
        return myjsonify(nodb_msg),500
    out={}
    sql=f"SELECT DISTINCT {colnam} FROM {tabnam} ORDER BY 1"
    items,columns,total_count,e=sql_select(dbengine,sql,with_total_count=False)
    if isinstance(e,str) and e=="ok":
        dbg("distinctvalues sql_select ok")
    else:
        dbg("distinctvalues sql_select error %s",str(e))
    if last_stmt_has_errors(e,out):
        return myjsonify(out),500
    out["data"]=[d[colnam] for d in pre_jsonify_items_transformer(items)]
    out["columns"]=columns
    out["total_count"]=len(items)
    dbg("leaving distinctvalues and return json result")
    dbg("out=%s",str(out))
    return myjsonify(out)

@api.route(api_root+'/exec/<db>/<procname>', methods=['POST'])
@token_required
def dbexec(tokdata,db,procname):
    """
    run execute procedure in database
    currently only for MSSQL
    ---
    tags:
      - Utils
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      -  name: body
         in: body
         description: all elements in body will be interpredid
         schema:
            required:
              - param_key
              - param_value
            properties:
              param_key:
                type: string
                description: parameter name
                example: " "
              param_value:
                type: string
                description: parameter value
      - name: db
        in: path
        type: string
        required: true
        description: id or alias of the database connection defined in repository table plainbi_datasource (0=repository)
      - name: procname
        in: path
        type: string
        required: true
        description: name of the stored procedure in the database
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
      500:
        description: Error running the exec procedure sql
        examples:
          application/json: 
            error: dbexec
            message: error bei dbexec <sql code>

    """
    global nodb_msg
    dbg("++++++++++ entering dbexec")
    dbg_api_call(request)
    audit(tokdata,request)
    dbengine=get_db_by_id_or_alias(db)
    if dbengine is None:
        return myjsonify(nodb_msg),500
    out={}
    dbtype=get_db_type(dbengine)

    data_bytes = request.get_data()
    sqlstmt = None # init
    data_string = data_bytes.decode('utf-8')
    item = json.loads(data_string.strip("'"))
    if dbtype=="mssql":
        sqlstmt = f"EXEC {procname}"
    else:
        out["error"] = "dbexec"
        out["message"] = "database type not supported"
        return myjsonify(out), 500

    first_key=True
    for key, value in item.items():
        if first_key:
            first_key = False
            sqlstmt += " "
        else:
            sqlstmt += ", "
        sqlstmt += f"{key} = {value}"
    dbg("dbexec: sql=%s",sqlstmt)

    try:
        #items, columns = db_exec(dbengine,sqlstmt)
        x = db_exec(dbengine,sqlstmt)
        dbg(str(type(x)))
        if isinstance(x,tuple):
            out["data"]=x[0]
            out["columns"]=x[1]
        else:
            out["message"]="sql executed"
    except SQLAlchemyError as e_sqlalchemy:
        err("dbexec_sql_errors: %s", str(e_sqlalchemy))
        if last_stmt_has_errors(e_sqlalchemy, out):
            out["error"]+="-dbexec"
            out["message"]+=" bei dbexec"
        return myjsonify(out), 500
    except Exception as e:
        err("dbexec exception: %s ",str(e))
        if last_stmt_has_errors(e, out):
            out["error"]+="-dbexec"
            out["message"]+=" beim dbexec"
        return myjsonify(out), 500

    if isinstance(out,dict):
        if "error" in out.keys():
            return myjsonify(out), 500
    return myjsonify(out)

###########################
##
## CRUD
##
###########################

# Define routes for CRUD operations
@api.route(api_prefix+'/<db>/<tab>', methods=['GET'])
@token_required
def get_all_items(tokdata,db,tab):
    """
    get database table contents (all rows)

    returns json with keys "data", "columns", "total_count"

    ---
    tags:
      - CRUD
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: db
        in: path
        type: string
        required: true
        description: id or alias of the database connection defined in repository table plainbi_datasource (0=repository)
      - name: tab
        in: path
        type: string
        required: true
        description: name of table in database 
      - name: v
        in: query
        type: boolean
        allowEmptyValue: true
        description: versions enabled 
      - name: cols
        in: query
        type: string
        required: false
        description: comma separated list of columns to get
      - name: q
        in: query
        type: string
        description: filter condition over all columns. if separated by blanks conditions will be connected with AND over all columns
      - name: filter
        in: query
        type: string
        description: a comma separated list of filter condition in the form column:value to search in individual columns. "~" instead of ":" means LIKE %value%, "!" means not equal
      - name: offset
        in: query
        type: integer
        description: start with row <offset> (for pagination)
      - name: limit
        in: query
        type: integer
        description: maximum number of rows to return  (for pagination)
      - name: order_by
        in: query
        type: string
        description: order by clause
      - name: customsql
        in: query
        type: string
        description: id or alias of sql in repository table plainbi_customersql. This replaces the tablename, bei "!" not equal
      - name: format
        in: query
        type: string
        description: output format XLSX/CSV/TXT
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering get_all_items")
    dbg_api_call(request)
    dbg("get_all_items: param tab is <%s>",str(tab))
    audit(tokdata,request)
    dbengine=get_db_by_id_or_alias(db)
    if dbengine is None:
        return myjsonify(nodb_msg),500
    out={}
    cols=request.args.get('cols')
    is_versioned = True if request.args.get('v') is not None else False
    myfilter, out = parse_filter(request.args.get('q'),request.args.get('filter'), out)
    if "error" in out.keys():
        return myjsonify(out), 500
    offset = request.args.get('offset')
    limit = request.args.get('limit')
    order_by = request.args.get('order_by')
    mycustomsql = request.args.get('customsql')
    dbg("pagination offset=%s limit=%s",offset,limit)
    items,columns,total_count,e=sql_select(dbengine,tab,order_by,offset,limit,with_total_count=True,versioned=is_versioned,filter=myfilter,customsql=mycustomsql,column_list=cols)
    if isinstance(e,str) and e=="ok":
        dbg("get_all_items sql_select ok")
    else:
        dbg("get_all_items sql_select error %s",str(e))
    if last_stmt_has_errors(e,out):
        return myjsonify(out),500
    has_format_param = True if request.args.get('format') is not None else False
    if has_format_param: # we want to download the data in CSV or Excel Format
        dbg("get_all_items: download data")
        fmt=request.args.get('format')
        df = pd.DataFrame(items)
        if fmt=="XLSX":
            dbg("get_all_items: XLSX format")
            tmpfile=os.path.join(tempfile.gettempdir(),'mydata'+datetime.now().strftime("%Y%m%d_%H%M%S")+'.xlsx')
            try:
                output = pd.ExcelWriter(tmpfile,engine="xlsxwriter")
                output.book.set_properties({"encoding":"utf-8"})
                fmt_xl.header_style = None
                df.to_excel(output, index=False, sheet_name="daten")
                output.close()
            except Exception as e0:
                err("get_all_items to_excel exception: %s ",str(e0))
                out["error"]="get_all_items-toxls"
                out["message"]="Fehler beim Prozessieren der Daten für den Download (XLSX)"
                out["detail"]=str(e0)
                err(traceback.format_exc())
                log.exception(e0)
                return myjsonify(out), 500
            with open(tmpfile, 'rb') as file:
                response = Response(
                    file.read(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment;filename=mydata.xlsx'}
                )
                dbg("get_all_items: return response")
                dbg(response)
                return response
        elif fmt=="CSV":
            dbg("get_all_items: CSV format")
            tmpfile=os.path.join(tempfile.gettempdir(),'mydata'+datetime.now().strftime("%Y%m%d_%H%M%S")+'.csv')
            # Prepare the CSV file
            try:
                df.to_csv(tmpfile, index=False)
            except Exception as e0:
                err("get_all_items to_csv exception: %s ",str(e0))
                out["error"]="get_all_items-tocsv"
                out["message"]="Fehler beim Prozessieren der Daten für den Download (CSV)"
                out["detail"]=str(e0)
                err(traceback.format_exc())
                log.exception(e0)
                return myjsonify(out), 500
            # Return the Excel file as a download
            with open(tmpfile, 'rb') as file:
                response = Response(
                    file.read(),
                    mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment;filename=mydata.csv'}
                )
                dbg(response)
                return response
        elif fmt=="TXT":
            dbg("get_all_items txt separated with tabs")
            tmpfile=os.path.join(tempfile.gettempdir(),'mydata'+datetime.now().strftime("%Y%m%d_%H%M%S")+'.txt')
            df.to_csv(tmpfile, index=False, sep='\t', quoting=csv.QUOTE_NONE)
            # Return the Excel file as a download
            with open(tmpfile, 'rb') as file:
                response = Response(
                    file.read(),
                    mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment;filename=mydata.csv'}
                )
                dbg(response)
                return response
        else: 
            out["error"]="get_all_items-invalid-format"
            out["message"]="Das Format muss XLSX/CSV/TXT sein"
            out["detail"]=None
            return myjsonify(out), 500
    out["data"]=pre_jsonify_items_transformer(items)
    out["columns"]=columns
    out["total_count"]=total_count
    dbg("leaving get_all_items and return json result")
    dbg("out=%s",str(out))
    return myjsonify(out)

# Define routes for CRUD operations

@api.route(api_prefix+'/<db>/<tab>/<pk>', methods=['GET'])
@token_required
def get_item(tokdata,db,tab,pk):
    """
    get a specific row from a table given by database tablename and id (or any primary key)

    returns jsons with key "data"  

    ---
    tags:
      - CRUD
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: db
        in: path
        type: string
        required: true
        description: id or alias of the database connection defined in repository table plainbi_datasource (0=repository)
      - name: tab
        in: path
        type: string
        required: true
        description: name of table in database 
      - name: pk
        in: path
        type: string
        required: true
        description: value of the primary key for the row to get  if more then one column in pk then comma separated. Values can be transferred url-safe-base64-encoded when string is in form [base64@<base64urlsafeencodedstring>]
      - name: cols
        in: query
        type: string
        required: false
        description: comma separated list of columns to get
      - name: pk
        in: query
        type: string
        description: column name of pk if it cant be extracted from metadata. (or comma separated list of columns if pk is combined)
      - name: v
        in: query
        type: boolean
        allowEmptyValue: true
        description: versioning enabled 
      - name: customsql
        in: query
        type: string
        description: id or alias of sql in repository table plainbi_customersql. This replaces the tablename 

    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering get_item")
    dbg_api_call(request)
    dbg("get_items: param tab is <%s>",str(tab))
    dbg("get_items: param pk/id is <%s>",str(pk))
    dbengine=get_db_by_id_or_alias(db)
    if dbengine is None:
        return myjsonify(nodb_msg),500
    audit(tokdata,request)
    # check options
    out={}
    is_versioned=False
    pkcols=[]
    cols=None
    if len(request.args) > 0:
        for key, value in request.args.items():
            dbg("arg: %s val: %s",key,value)
            if key=="pk":
                pkcols=value.split(",")
                dbg("pk option %s",pkcols)
            if key=="cols":
                cols=value
                dbg("cols option %s",cols)
            if key=="v":
                is_versioned=True
                dbg("versions enabled")
    mycustomsql = request.args.get('customsql')
    dbg("tab %s pk %s")
    # check if pk is compound
    if pk == '#' or pk == '@':
        data_bytes = request.get_data()
        dbg("get_item: data")
        dbg("get_item: databytes: %s",data_bytes)
        data_string = data_bytes.decode('utf-8')
        dbg("datastring: %s",data_string)
        pk = json.loads(data_string.strip("'"))
    else:
        pk=prep_pk_from_url(pk)

    #
    out=get_item_raw(dbengine, tab, pk, pk_column_list=pkcols, column_list=cols, versioned=is_versioned, customsql=mycustomsql)
    if "data" in out.keys():
        if len(out["data"])>0:
            # jk20240910 for date formatting on output
            pre_jsonify_items_transformer(out["data"])
            dbg("out:%s",str(out))
            dbg("leaving get_item with success and json result")
            try:
                json_out = jsonify(out)
            except Exception as ej:
                err("get_item: jsonify Error: %s",str(ej))
                log.exception(ej)
                return "get_item: jsonify Error",500
            return myjsonify(out)
            #return Response(jsonify(out),status=204)
        else:
            dbg("no record found")
            # return Response(status=204)
            dbg("leaving get_item with 204 no record forund")
            return ("kein datensatz gefunden",204,"")
    # return (resp.text, resp.status_code, resp.headers.items())
    dbg("leaving get_item with error 500 and return json result")
    return myjsonify(out),500

@api.route(api_prefix+'/<db>/<tab>/<pk>', methods=['POST'])
@token_required
def get_item_post(tokdata,db,tab,pk):
    """
    get a specific row from a table given by database tablename and id (or any primary key)

    returns jsons with key "data"  

    ---
    tags:
      - CRUD
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: db
        in: path
        type: string
        required: true
        description: id or alias of the database connection defined in repository table plainbi_datasource (0=repository)
      - name: tab
        in: path
        type: string
        required: true
        description: name of table in database 
      - name: pk
        in: path
        type: string
        required: true
        description: value of the primary key for the row to get  if pk="#" or pk="@" then pk is taken request.data    if more then on column in pk then comma separated
      - name: cols
        in: query
        type: string
        required: false
        description: comma separated list of columns to get
      - name: v
        in: query
        type: boolean
        allowEmptyValue: true
        description: versioning enabled 
      - name: customsql
        in: query
        type: string
        description: id or alias of sql in repository table plainbi_customersql. This replaces the tablename 
      - name: body
        in: body
        type: string
        required: true
        description: pk in request body instead of url

    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            columns: id,...
            data: x,y,...
            total_count: 1
    """
    dbg("++++++++++ entering get_item_post")
    dbg_api_call(request)
    dbg("get_items: param tab is <%s>",str(tab))
    dbg("get_items: param pk/id is <%s>",str(pk))
    dbengine=get_db_by_id_or_alias(db)
    if dbengine is None:
        return myjsonify(nodb_msg),500
    audit(tokdata,request)
    # check options
    out={}
    is_versioned=False
    pkcols=[]
    cols=None
    if len(request.args) > 0:
        for key, value in request.args.items():
            dbg("arg: %s val: %s",key,value)
            if key=="pk":
                pkcols=value.split(",")
                dbg("pk option %s",pkcols)
            if key=="cols":
                cols=value
                dbg("cols option %s",cols)
            if key=="v":
                is_versioned=True
                dbg("versions enabled")
    mycustomsql = request.args.get('customsql')
    dbg("tab %s pk %s")
    # check if pk is compound
    if pk == '#' or pk == '@':
        data_bytes = request.get_data()
        dbg("get_item: data")
        dbg("get_item: databytes: %s",data_bytes)
        data_string = data_bytes.decode('utf-8')
        dbg("datastring: %s",data_string)
        pk = json.loads(data_string.strip("'"))
    else:
        pk=prep_pk_from_url(pk)

    #
    out=get_item_raw(dbengine, tab, pk, pk_column_list=pkcols, column_list=cols, versioned=is_versioned, customsql=mycustomsql)
    if "data" in out.keys():
        if len(out["data"])>0:
            dbg("get_item_raw call output in get_item_post:"+str(out))
            pre_jsonify_items_transformer(out["data"])
            dbg("out:%s",str(out))
            dbg("leaving get_item with success and json result")
            try:
                json_out = jsonify(out)
            except Exception as ej:
                err("get_item: jsonify Error: %s",str(ej))
                log.exception(ej)
                return "get_item: jsonify Error",500
            return myjsonify(out)
            #return Response(jsonify(out),status=204)
        else:
            dbg("no record found")
            # return Response(status=204)
            dbg("leaving get_item with 204 no record forund")
            return ("kein datensatz gefunden",204,"")
    # return (resp.text, resp.status_code, resp.headers.items())
    dbg("leaving get_item with error 500 and return json result")
    return myjsonify(out),500


@api.route(api_prefix+'/<db>/<tab>', methods=['POST'])
@token_required
def create_item(tokdata,db,tab):
    """
    create a new row in the database (insert)

    returns json mit den keys "data"  i.e. the inserted row (might have new data f.e. sequence values, trigger)

    ---
    tags:
      - CRUD
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: db
        in: path
        type: string
        required: true
        description: id or alias of the database connection defined in repository table plainbi_datasource (0=repository)
      - name: tab
        in: path
        type: string
        required: true
        description: name of table in database 
      - name: v
        in: query
        type: boolean
        allowEmptyValue: true
        description: versions enabled 
      - name: usercol
        type: string
        required: false
        description: name of the column which should be filled with the username
      - name: pk
        in: query
        type: string
        description: column name of pk if it cant be extracted from metadata. (or comma separated list of columns if pk is combined)
      - name: seq
        in: query
        type: string
        description: name of a database sequence to create a new primary key value when inserting the row
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering create_item")
    dbg_api_call(request)
    dbg("create_item: param tab is <%s>",str(tab))
    audit(tokdata,request)
    dbengine=get_db_by_id_or_alias(db)
    if dbengine is None:
        return myjsonify(nodb_msg),500
    out={}
    pkcols=[]
    is_versioned=False
    seq=None
    usercol=None
    # check options
    if len(request.args) > 0:
        for key, value in request.args.items():
            dbg("arg: %s val: %s",key,value)
            if key=="pk":
                pkcols=value.split(",")
                dbg("pk option %s",pkcols)
            if key=="seq":
                seq=value
                dbg("pk sequence %s",seq)
            if key=="usercol":
                usercol=value
                dbg("usercol %s",usercol)
            if key=="v":
                is_versioned=True
                dbg("versions enabled")
    dbg("create_item tab %s pkcols %s seq %s",tab,pkcols,seq)
    mycustomsql = request.args.get('customsql')

    data_bytes = request.get_data()
    dbg("create_item 7")
    dbg("databytes: %s",data_bytes)
    data_string = data_bytes.decode('utf-8')
    dbg("datastring: %s",data_string)
    item = json.loads(data_string.strip("'"))
    if usercol is not None:
        item[usercol]=tokdata['username']
        dbg("usercol %s set to %s",usercol,item[usercol])
    out = db_ins(dbengine,tab,item,pkcols,is_versioned,seq,changed_by=tokdata['username'],customsql=mycustomsql)
    if isinstance(out,dict):
        if "error" in out.keys():
            return myjsonify(out), 400
    return myjsonify(out)


@api.route(api_prefix+'/<db>/<tab>/<pk>', methods=['PUT'])
@token_required
def update_item(tokdata,db,tab,pk):
    """
    update a row in a table

    ---
    tags:
      - CRUD
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: db
        in: path
        type: string
        required: true
        description: id or alias of the database connection defined in repository table plainbi_datasource (0=repository)
      - name: tab
        in: path
        type: string
        required: true
        description: name of table in database 
      - name: pk
        in: path
        type: string
        required: true
        description: value of the primary key for the row to get  if pk=# then pk is taken request.data    if more then on column in pk then comma separated
                     If a value is in form [base64@<base64urlsafeencodedstring>] then it is url-safe base64 encoded
      - name: pk
        in: query
        type: string
        description: column name of pk if it cant be extracted from metadata. (or comma separated list of columns if pk is combined)
      - name: v
        in: query
        type: boolean
        allowEmptyValue: true
        description: versioning enabled 
      - name: usercol
        type: string
        required: false
        description: name of the column which should be filled with the username
      - name: body
        in: body
        required: true
        schema:
           required:
             - feld
           properties:
             feld:
               type: string
               description: Feld Inhalt
               example: "fekd"
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering update_item")
    dbg_api_call(request)
    dbg("update_item: param tab is <%s>",str(tab))
    dbg("update_item: param pk is <%s>",str(pk))
    audit(tokdata,request)
    dbengine=get_db_by_id_or_alias(db)
    if dbengine is None:
        return myjsonify(nodb_msg),500
    out={}
    pkcols=[]
    is_versioned=False
    usercol=None
    # check options
    if len(request.args) > 0:
        for key, value in request.args.items():
            dbg("arg: %s val: %s",key,value)
            if key=="pk":
                pkcols=value.split(",")
                dbg("pk option %s",pkcols)
            if key=="v":
                is_versioned=True
                dbg("versions enabled")
            if key=="usercol":
                usercol=value
                dbg("usercol enabled for col %s",usercol)
    mycustomsql = request.args.get('customsql')
    # check if pk is compound
    pk=prep_pk_from_url(pk)
    # check pk from compound key 
    if len(pkcols)==0:
        # pk columns are not explicitly given as url parameter
        if isinstance(pk,dict):
           # there is an url pk in form (col:val)
           pkcols=list(pk.keys())
           dbg("pk columns from url form (col:val[:col2:val2...])")
    else:
        dbg("pk columns explicitly from url parameter")
    #
    data_bytes = request.get_data()
    dbg("databytes: %s",data_bytes,dbglevel=3)
    data_string = data_bytes.decode('utf-8')
    dbg("datastring: %s",data_string,dbglevel=3)
    item = json.loads(data_string.strip("'"))
    #item = {key: request.data[key] for key in request.data}
    dbg("item %s",item,dbglevel=3)
    if usercol is not None:
        item[usercol]=tokdata['username']
        dbg("usercol %s set to %s",usercol,item[usercol])

    
    out = db_upd(dbengine, tab, pk, item, pkcols, is_versioned, changed_by=tokdata['username'], customsql=mycustomsql)
    if isinstance(out,dict):
        if "error" in out.keys():
            err("=update_item out error (see stdout for more) ======================")
            print("==============================================")
            print("=update_item out error================================")
            pprint.pprint(out)
            print("==============================================")
            return myjsonify(out), 400

    return myjsonify(out)

@api.route(api_prefix+'/<db>/<tab>/<pk>', methods=['DELETE'])
@token_required
def delete_item(tokdata,db,tab,pk):
    """
    delete a row in a database

    returns 200 or json with error msg

    ---
    tags:
      - CRUD
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: db
        in: path
        type: string
        required: true
        description: id or alias of the database connection defined in repository table plainbi_datasource (0=repository)
      - name: tab
        in: path
        type: string
        required: true
        description: name of table in database 
      - name: pk
        in: path
        type: string
        required: true
        description: value of the primary key for the row to get  if pk=# then pk is taken request.data    if more then on column in pk then comma separated. 
                     If a value is in form [base64@<base64urlsafeencodedstring>] then it is url-safe base64 encoded
      - name: pk
        in: query
        type: string
        description: column name of pk if it cant be extracted from metadata. (or comma separated list of columns if pk is combined)
      - name: v
        in: query
        type: boolean
        allowEmptyValue: true
        description: versioning enabled 
      - name: usercol
        type: string
        required: false
        description: name of the column which should be filled with the username
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering delete_item")
    dbg_api_call(request)
    dbg("delete_item: param tab is <%s>",str(tab))
    dbg("delete_item: param pk is <%s>",str(pk))
    audit(tokdata,request)
    dbengine=get_db_by_id_or_alias(db)
    if dbengine is None:
        return myjsonify(nodb_msg),500
    out={}
    pkcols=[]
    is_versioned=False
    usercol=None
    # check options
    if len(request.args) > 0:
        for key, value in request.args.items():
            dbg("arg: %s val: %s",key,value)
            if key=="pk":
                pkcols=value.split(",")
                dbg("pk option %s",pkcols)
            if key=="v":
                is_versioned=True
                dbg("versions enabled")
            if key=="usercol":
                usercol=value
                dbg("usercol enabled for col %s",usercol)
    dbg("delete_item tab %s pkcols %s ",tab,pkcols)

    pk=prep_pk_from_url(pk)
    dbg("delete_item tab %s pk %s",tab,pk)
    # check pk from compound key 
    if len(pkcols)==0:
        # pk columns are not explicitly given as url parameter
        if isinstance(pk,dict):
           # there is an url pk in form (col:val)
           pkcols=list(pk.keys())
           dbg("pk columns from url form (col:val[:col2:val2...])")
    else:
        dbg("pk columns explicitly from url parameter")

    dbg("############# pk columns for delete is %s",str(pkcols))
    out = db_del(dbengine, tab, pk, pkcols, is_versioned, changed_by=tokdata['username'])
    if isinstance(out,dict):
        if "error" not in out.keys():
            return 'Record deleted successfully', 200
        else:
            return myjsonify(out), 400
    return myjsonify(out)


@api.route(api_metadata_prefix+'/<db>/tables', methods=['GET'])
@token_required
def get_metadata_tables(tokdata,db):
    """
    get names of all accessible tables in the database

    returns json with key "data"  

    ---
    tags:
      - Metadata
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: db
        in: path
        type: string
        required: true
        description: id or alias of the database connection defined in repository table plainbi_datasource (0=repository)
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering get_metadata_tables")
    dbg_api_call(request)
    audit(tokdata,request)
    dbengine=get_db_by_id_or_alias(db)
    if dbengine is None:
        return myjsonify(nodb_msg),500
    offset = request.args.get('offset')
    limit = request.args.get('limit')
    order_by = request.args.get('order_by')
    out={}
    items,columns,total_count,e=sql_select(dbengine,metadata_tab_query,order_by,offset,limit,with_total_count=False)
    dbg("get_metadata_tables sql_select error %s",str(e))
    if last_stmt_has_errors(e,out):
        return myjsonify(out),500
    out["data"]=pre_jsonify_items_transformer(items)
    out["columns"]=columns
    out["total_count"]=total_count
    return myjsonify(out)

@api.route(api_metadata_prefix+'/<db>/table/<tab>', methods=['GET'])
@token_required
def get_metadata_tab_columns(tokdata,db,tab):
    """
    get metadata of a table from the database dictionary

    returns json with columns and datatypes

    ---
    tags:
      - Metadata
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: db
        in: path
        type: string
        required: true
        description: id or alias of the database connection defined in repository table plainbi_datasource (0=repository)
      - name: tab
        in: path
        type: string
        required: true
        description: name of table in database 
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering get_metadata_tab_columns")
    dbg_api_call(request)
    dbg("get_metadata_tab_columns: param tab is <%s>",str(tab))
    audit(tokdata,request)
    dbengine=get_db_by_id_or_alias(db)
    if dbengine is None:
        return myjsonify(nodb_msg),500
    out={}
    pkcols=None
    if len(request.args) > 0:
        for key, value in request.args.items():
            dbg("arg: %s val: %s",key,value)
            if key=="pk":
                pkcols=value.split(",")
                dbg("pk option %s",pkcols)
    dbg('get_metadata_tab_columns: for %s',tab)
    try:
        metadata=get_metadata_raw(dbengine,tab,pk_column_list=pkcols)
    except SQLAlchemyError as e_sqlalchemy:
        if last_stmt_has_errors(e_sqlalchemy, out):
            out["error"]+="-get_metadata_tab_columns"
            out["message"]+=" beim Lesen der Tabellen Metadaten"
        return myjsonify(out),500
    except Exception as e:
        if last_stmt_has_errors(e, out):
            out["error"]+="-get_metadata_tab_columns"
            out["message"]+=" beim Lesen der Tabellen Metadaten"
        return myjsonify(out),500
    return myjsonify(metadata)

###########################
##
## REPO
##
###########################

# Define routes for REPO operations
@api.route(repo_api_prefix+'/resources', methods=['GET'])
@token_required
def get_resource(tokdata):
    """
    get the resources from the repository

    returns json of all applications, adhocs, and external resources

    ---
    tags:
      - Repo
    security:
    - APIKeyHeader: ['Authorization']
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering get_resource")
    dbg_api_call(request)
    audit(tokdata,request)
    prof=get_profile(config.repoengine,tokdata['username'])
    user_id=prof["user_id"]
    out={}
    offset = request.args.get('offset')
    limit = request.args.get('limit')
    order_by = request.args.get('order_by')
    dbg("pagination offset=%s limit=%s",offset,limit)
    
    w_app=add_auth_to_where_clause("plainbi_application",None,user_id)
    w_adhoc=add_auth_to_where_clause("plainbi_adhoc",None,user_id)
    w_ext_res=add_auth_to_where_clause("plainbi_external_resource",None,user_id)
    if not hasattr(config,"repo_db_type"):
        config.repo_db_type=get_db_type(config.repoengine)
    dbg("get_resource config.repo_db_type=%s",config.repo_db_type)
    if config.repo_db_type == 'mssql':
        concat_op='+'
    else:
        concat_op='||'
    dbg("get_resource concat_op=%s",concat_op)
    
    resource_sql=f"""select
'application_'{concat_op}cast(id as varchar) as id
, name
, '/apps/'{concat_op}alias as url
, '_self' as target
, null as output_format
, null as description
, null as source
, null as dataset
, 'application' as resource_type
, 'Applikation' as resource_type_de
from plainbi_application pa
{w_app}
union all
select
'adhoc_'{concat_op}cast(id as varchar) as id
, name
, '/adhoc/' {concat_op} cast(id as varchar) {concat_op} case when coalesce(output_format, 'HTML') <> 'HTML' then '?format='{concat_op}output_format else '' end as url
, '_self' as target
, coalesce(output_format, 'HTML') output_format
, description
, 'Adhoc' as source
, null as dataset
, 'adhoc' as resource_type
, 'Adhoc' as resource_type_de
from plainbi_adhoc padh
{w_adhoc}
union all
select
'external_resource_'{concat_op}cast(id as varchar) as id
, name
, url
, '_blank' as target
, null as output_format
, description
, source
, dataset
, 'external_resource' as resource_type
, source as resource_type_de
from plainbi_external_resource per
{w_ext_res}
"""
    items,columns,total_count,e=sql_select(config.repoengine,resource_sql,order_by,offset,limit,with_total_count=True,is_repo=True,user_id=prof["user_id"])
    dbg("get_resource sql_select error %s",str(e))
    if last_stmt_has_errors(e,out):
        return myjsonify(out),500
    out["data"]=pre_jsonify_items_transformer(items)
    out["columns"]=columns
    out["total_count"]=total_count
    return myjsonify(out)


# mir zugeordnete Gruppen
@api.route(repo_api_prefix+'/groups', methods=['GET'])
@token_required
def get_my_groups(tokdata):
    """
    get my groups
    ---
    tags:
      - Repo
    security:
    - APIKeyHeader: ['Authorization']
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering get_my_groups")
    dbg_api_call(request)
    audit(tokdata,request)
    prof=get_profile(config.repoengine,tokdata['username'])
    user_id=prof["user_id"]
    out={}
    if not hasattr(config,"repo_db_type"):
        config.repo_db_type=get_db_type(config.repoengine)
    #mysql="select g.id, g.name from plainbi_user_to_group ug join plainbi_group g on ug.group_id = g.id where ug.user_id="+prof["user_id"]
    mysql=f"select distinct g.id, g.name from plainbi_user_to_group ug join plainbi_group g on ug.group_id = g.id where ug.user_id={user_id} or {user_id} in (select id from plainbi_user where role_id=1)"
    dbg("get_my_groups sql: %s",mysql)
    items,columns,total_count,e=sql_select(config.repoengine,mysql,order_by=None,offset=None,limit=None,with_total_count=True,is_repo=True,user_id=prof["user_id"])
    dbg("get_my_groups sql_select error %s",str(e))
    if last_stmt_has_errors(e,out):
        return myjsonify(out),500
    out["data"]=pre_jsonify_items_transformer(items)
    out["columns"]=columns
    out["total_count"]=total_count
    return myjsonify(out)

# 
@api.route(repo_api_prefix+'/group/<gid>/resources', methods=['GET'])
@token_required
def get_group_resources(tokdata,gid):
    """
    Resourcen gefiltert auf die Gruppe
    ---
    tags:
      - Repo
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: gid
        in: path
        type: string
        required: true
        description: group id  (or "nogroup" for all resoures not in a group (admins only))
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering get_my_groups")
    dbg_api_call(request)
    audit(tokdata,request)
    prof=get_profile(config.repoengine,tokdata['username'])
    user_id=prof["user_id"]
    user_is_admin_flag=prof["user_is_admin"]
    out={}
    if not hasattr(config,"repo_db_type"):
        config.repo_db_type=get_db_type(config.repoengine)
    dbg("get_resource config.repo_db_type=%s",config.repo_db_type)
    if config.repo_db_type == 'mssql':
        concat_op='+'
    else:
        concat_op='||'
    dbg("get_resource concat_op=%s",concat_op)
    if gid=="nogroup":
        resource_sql=f"""select
    'application_'{concat_op}cast(id as varchar) as id
    , name
    , '/apps/'{concat_op}alias as url
    , '_self' as target
    , null as output_format
    , null as description
    , null as source
    , null as dataset
    , 'application' as resource_type
    , 'Applikation' as resource_type_de
    from plainbi_application pa
    where '{user_is_admin_flag}'='Y'
    and pa.id not in (select application_id from plainbi_application_to_group)
    union all
    select
    'adhoc_'{concat_op}cast(id as varchar) as id
    , name
    , '/adhoc/' {concat_op} cast(id as varchar) {concat_op} case when coalesce(output_format, 'HTML') <> 'HTML' then '?format='{concat_op}output_format else '' end as url
    , '_self' as target
    , coalesce(output_format, 'HTML') output_format
    , description
    , 'Adhoc' as source
    , null as dataset
    , 'adhoc' as resource_type
    , 'Adhoc' as resource_type_de
    from plainbi_adhoc padh
    where '{user_is_admin_flag}'='Y'
    and padh.id not in (select adhoc_id from plainbi_adhoc_to_group)
    union all
    select
    'external_resource_'{concat_op}cast(id as varchar) as id
    , name
    , url
    , '_blank' as target
    , null as output_format
    , description
    , source
    , dataset
    , 'external_resource' as resource_type
    , source as resource_type_de
    from plainbi_external_resource per
    where '{user_is_admin_flag}'='Y'
    and per.id not in (select external_resource_id from plainbi_external_resource_to_group)
    """
    else:
        if not is_id(gid):
            items, columns = db_exec(config.repoengine,f"select id from plainbi_group where alias='{gid}'")
            if len(items) > 0:
                gid=items[0]["id"]
            else:
                out["error"]="no-such-group-alias"
                out["message"]=f"Berechtigungsgruppe mit dem alias {gid} nicht gefunden"
                return myjsonify(out), 500
        else:
            items, columns = db_exec(config.repoengine,f"select id from plainbi_group where id={gid}")
            if len(items) < 1:
                out["error"]="no-such-group-id"
                out["message"]=f"Berechtigungsgruppe mit der ID {gid} nicht gefunden"
                return myjsonify(out), 500
        resource_sql=f"""select
    'application_'{concat_op}cast(id as varchar) as id
    , name
    , '/apps/'{concat_op}alias as url
    , '_self' as target
    , null as output_format
    , null as description
    , null as source
    , null as dataset
    , 'application' as resource_type
    , 'Applikation' as resource_type_de
    from plainbi_application pa
    join plainbi_application_to_group ag
    on pa.id=ag.application_id
    and ag.group_id={gid}
    and ag.group_id in (select ug.group_id from plainbi_user_to_group ug where ug.user_id={user_id} or {user_id} in (select id from plainbi_user where role_id=1))
    union all
    select
    'adhoc_'{concat_op}cast(id as varchar) as id
    , name
    , '/adhoc/' {concat_op} cast(id as varchar) {concat_op} case when coalesce(output_format, 'HTML') <> 'HTML' then '?format='{concat_op}output_format else '' end as url
    , '_self' as target
    , coalesce(output_format, 'HTML') output_format
    , description
    , 'Adhoc' as source
    , null as dataset
    , 'adhoc' as resource_type
    , 'Adhoc' as resource_type_de
    from plainbi_adhoc padh
    join plainbi_adhoc_to_group ag
    on padh.id = ag.adhoc_id
    and ag.group_id={gid}
    and ag.group_id in (select ug.group_id from plainbi_user_to_group ug where ug.user_id={user_id} or {user_id} in (select id from plainbi_user where role_id=1))
    union all
    select
    'external_resource_'{concat_op}cast(id as varchar) as id
    , name
    , url
    , '_blank' as target
    , null as output_format
    , description
    , source
    , dataset
    , 'external_resource' as resource_type
    , source as resource_type_de
    from plainbi_external_resource per
    join plainbi_external_resource_to_group rg
    on per.id=rg.external_resource_id
    and rg.group_id={gid}
    and rg.group_id in (select ug.group_id from plainbi_user_to_group ug where ug.user_id={user_id} or {user_id} in (select id from plainbi_user where role_id=1))
    """
    dbg("get_group_resources sql: %s",resource_sql)
    items,columns,total_count,e=sql_select(config.repoengine,resource_sql,order_by=None,offset=None,limit=None,with_total_count=True,is_repo=True,user_id=prof["user_id"])
    dbg("get_group_resources sql_select error %s",str(e))
    if last_stmt_has_errors(e,out):
        return myjsonify(out),500
    out["data"]=pre_jsonify_items_transformer(items)
    out["columns"]=columns
    out["total_count"]=total_count
    return myjsonify(out)


# Define routes for REPO operations
@api.route(repo_api_prefix+'/<tab>', methods=['GET'])
@token_required
def get_all_repos(tokdata,tab):
    """
    get table contents of table <tab> in the repository (table name without prefix plainbi_)

    returns json with keys "data", "columns", "total_count"

    ---
    tags:
      - Repo
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: tab
        in: path
        type: string
        required: true
        description: name of the repository table
      - name: q
        in: query
        type: string
        description: filter condition over all columns. if separated by blanks conditions will be connected with AND over all columns
      - name: filter
        in: query
        type: string
        description: a comma separated list of filter condition in the form column:value to search in individual columns. "~" instead of ":" means LIKE %value%, "!" means not equal
      - name: offset
        in: query
        type: integer
        description: start with row <offset> (for pagination)
      - name: limit
        in: query
        type: integer
        description: maximum number of rows to return  (for pagination)
      - name: order_by
        in: query
        type: string
        description: order by clause
      - name: customsql
        in: query
        type: string
        description: id or alias of sql in repository table plainbi_customersql. This replaces the tablename, bei "!" not equal
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering get_all_repos")
    dbg_api_call(request)
    dbg("get_all_repos: param tab is <%s>",str(tab))
    audit(tokdata,request)
    prof=get_profile(config.repoengine,tokdata['username'])
    out={}
    myfilter, out = parse_filter(request.args.get('q'),request.args.get('filter'), out)
    if "error" in out.keys():
        return myjsonify(out), 500
    offset = request.args.get('offset')
    limit = request.args.get('limit')
    order_by = request.args.get('order_by')
    mycustomsql = request.args.get('customsql')
    dbg("pagination offset=%s limit=%s",offset,limit)
    items,columns,total_count,e=sql_select(config.repoengine,repo_table_prefix+tab,order_by,offset,limit,filter=myfilter,with_total_count=True,is_repo=True,user_id=prof["user_id"],customsql=mycustomsql)
    dbg("get_all_repos sql_select error %s",str(e))
    if last_stmt_has_errors(e,out):
        return myjsonify(out),500
    out["data"]=pre_jsonify_items_transformer(items)
    out["columns"]=columns
    out["total_count"]=total_count
    return myjsonify(out)

@api.route(repo_api_prefix+'/<tab>/<pk>', methods=['GET'])
@token_required
def get_repo(tokdata,tab,pk):
    """
    get a specific row from a repository table

    returns json with keys "data"  

    ---
    tags:
      - Repo
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: tab
        in: path
        type: string
        required: true
        description: name of the repository table (without the plainbi_ prefix)
      - name: pk
        in: path
        type: string
        required: true
        description: primary key of the row to get from the repository table
      - name: pk
        in: query
        type: string
        description: column name of pk if it cant be extracted from metadata. (or comma separated list of columns if pk is combined)
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering get_repo")
    dbg("get_repo: param tab is <%s>",str(tab))
    dbg("get_repo: param pk is <%s>",str(pk))
    audit(tokdata,request)
    # check options
    prof=get_profile(config.repoengine,tokdata['username'])
    pkcols=[]
    if len(request.args) > 0:
        for key, value in request.args.items():
            dbg("arg: %s val: %s",key,value)
            if key=="pk":
                pkcols=value.split(",")
                dbg("pk option %s",pkcols)
    # check if pk is compound
    mycustomsql = request.args.get('customsql')
    pk=prep_pk_from_url(pk)
    if tab=="application" and (not is_id(pk)):
        # use alias
        out=get_item_raw(config.repoengine,repo_table_prefix+tab,pk,pk_column_list=["alias"],is_repo=True,user_id=prof["user_id"],customsql=mycustomsql)
    else:    
        out=get_item_raw(config.repoengine,repo_table_prefix+tab,pk,pk_column_list=pkcols,is_repo=True,user_id=prof["user_id"],customsql=mycustomsql)
    if "data" in out.keys():
        if len(out["data"])>0:
            #print("out:"+str(out))
            pre_jsonify_items_transformer(out["data"])
            #dbg("out:%s",str(out))
            dbg("return get_repo out:%s",str(out)[:255],dbglevel=3)
            return myjsonify(out)
            #return Response(jsonify(out),status=204)
        else:
            dbg("no record found")
            # return Response(status=204)
            return ("kein datensatz gefunden",204,"")
    # return (resp.text, resp.status_code, resp.headers.items())
    dbg("return get_repo but no data")
    return myjsonify(out),500


@api.route(repo_api_prefix+'/<tab>', methods=['POST'])
@token_required
def create_repo(tokdata,tab):
    """
    insert a new row into a repository table 

    Parameters
    tab : repository table name (without prefix plainbi_)
    
    Url Options:
        pk=
        seq=  Name of Sequence for PK, in case None/Null is sent

    return json with keys "data" of the newly inserted row

    ---
    tags:
      - Repo
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: tab
        in: path
        type: string
        required: true
        description: name of the repository table (without the plainbi_ prefix)
      - name: pk
        in: query
        type: string
        description: column name of pk if it cant be extracted from metadata. (or comma separated list of columns if pk is combined)
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering create_repo")
    dbg("create_repo: param tab is <%s>",str(tab))
    audit(tokdata,request)
    prof=get_profile(config.repoengine,tokdata['username'])
    out={}
    pkcols=[]
    is_versioned=False
    # check options
    dbg("create_repo: check url params")
    if len(request.args) > 0:
        for key, value in request.args.items():
            dbg("arg: %s val: %s",key,value)
            if key=="pk":
                pkcols=value.split(",")
                dbg("pk option %s",pkcols)
            if key=="v":
                is_versioned=True
                dbg("versions enabled")
    dbg("create_repo tab %s pkcols %s",tab,pkcols)
    mycustomsql = request.args.get('customsql')

    data_bytes = request.get_data()
    dbg("databytes: %s",data_bytes)
    data_string = data_bytes.decode('utf-8')
    dbg("datastring: %s",data_string)
    item = json.loads(data_string.strip("'"))
    db_typ = get_db_type(config.repoengine)
    if tab in ["adhoc","application","datasource","external_resource","group","lookup","role","user","group","customsql","adhoc_parameter"]:
        if db_typ=="sqlite":
           seq=tab
        elif db_typ in ("mssql","postgres","oracle"):
           seq="plainbi_"+tab+"_seq"
        else:
           err("create_repo: unknown repo database type")
           seq=None
    else:
        seq=None
    out = db_ins(config.repoengine,repo_table_prefix+tab,item,pkcols,is_versioned,seq,is_repo=True,customsql=mycustomsql)
    if isinstance(out,dict):
        if "error" in out.keys():
            return myjsonify(out), 400
    return myjsonify(out)


@api.route(repo_api_prefix+'/<tab>/<pk>', methods=['PUT'])
@token_required
def update_repo(tokdata,tab,pk):
    """
    update a row in the repository

    Parameters
    tab : repository table name (without prefix plainbi_)
    pk : Primary Key Identifier (Primary Key)
    
    Url Options:
        pk=

    returns json with keys "data" of the updated row

    ---
    tags:
      - Repo
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: tab
        in: path
        type: string
        required: true
        description: name of the repository table (without the plainbi_ prefix)
      - name: pk
        in: path
        type: string
        required: true
        description: primary key of the row to get from the repository table
      - name: pk
        in: query
        type: string
        description: column name of pk if it cant be extracted from metadata. (or comma separated list of columns if pk is combined)
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering update_repo")
    dbg_api_call(request)
    dbg("update_repo: param tab is <%s>",str(tab))
    dbg("update_repo: param pk is <%s>",str(pk))
    audit(tokdata,request)
    prof=get_profile(config.repoengine,tokdata['username'])
    out={}
    pkcols=[]
    is_versioned=False
    # check options
    if len(request.args) > 0:
        for key, value in request.args.items():
            dbg("arg: %s val: %s",key,value)
            if key=="pk":
                pkcols=value.split(",")
                dbg("pk option %s",pkcols)
    mycustomsql = request.args.get('customsql')
    # check if pk is compound
    pk=prep_pk_from_url(pk)
    # check pk from compound key 
    if len(pkcols)==0:
        # pk columns are not explicitly given as url parameter
        if isinstance(pk,dict):
           # there is an url pk in form (col:val)
           pkcols=list(pk.keys())
           dbg("pk columns from url form (col:val[:col2:val2...])")
    else:
        dbg("pk columns explicitly from url parameter")
    
    data_bytes = request.get_data()
    dbg("databytes: %s",data_bytes,dbglevel=3)
    data_string = data_bytes.decode('utf-8')
    dbg("datastring: %s",data_string,dbglevel=3)
    item = json.loads(data_string.strip("'"))
    dbg("datastring: %s",str(item),dbglevel=3)

    out = db_upd(config.repoengine,repo_table_prefix+tab,pk,item,pkcols,is_versioned,is_repo=True,customsql=mycustomsql)
    if isinstance(out,dict):
        if "error" in out.keys():
            err("=update_repo out error (see stdout for more) ======================")
            print("==============================================")
            print("=update_repo out error================================")
            pprint.pprint(out)
            print("==============================================")
            return myjsonify(out), 400
    return myjsonify(out)


@api.route(repo_api_prefix+'/<tab>/<pk>', methods=['DELETE'])
@token_required
def delete_repo(tokdata,tab,pk):
    """
    delete a row in the repositoy

    Parameters
    tab : repository table name (without prefix plainbi_)
    pk : Primary Key Identifier (Primary Key) of the row to be deleted
    
    Url Options:
        pk=

    returns 200 or json of error message

    ---
    tags:
      - Repo
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: tab
        in: path
        type: string
        required: true
        description: tablename in repository
      - name: pk
        in: path
        type: string
        required: true
        description: primary key identifiery of the row to delete in the repository table
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering delete_repo")
    dbg_api_call(request)
    dbg("delete_repo: param tab is <%s>",str(tab))
    dbg("delete_repo: param pk is <%s>",str(pk))
    audit(tokdata,request)
    prof=get_profile(config.repoengine,tokdata['username'])
    out={}
    pkcols=[]
    is_versioned=False
    # check options
    if len(request.args) > 0:
        for key, value in request.args.items():
            dbg("arg: %s val: %s",key,value)
            if key=="pk":
                pkcols=value.split(",")
                dbg("pk option %s",pkcols)
            if key=="v":
                is_versioned=True
                dbg("versions enabled")
    dbg("delete_item tab %s pkcols %s ",tab,pkcols)

    pk=prep_pk_from_url(pk)
    dbg("delete_repo tab %s pk %s",tab,pk)
    # check pk from compound key 
    if len(pkcols)==0:
        # pk columns are not explicitly given as url parameter
        if isinstance(pk,dict):
           # there is an url pk in form (col:val)
           pkcols=list(pk.keys())
           dbg("pk columns from url form (col:val[:col2:val2...])")
    else:
        dbg("pk columns explicitly from url parameter")

    dbg("############# pk columns for delete is %s",str(pkcols))
    out = db_del(config.repoengine,repo_table_prefix+tab,pk,pkcols,is_versioned,is_repo=True)
    if isinstance(out,dict):
        if "error" not in out.keys():
            return 'Repo Record deleted successfully', 200
        else:
            return myjsonify(out), 400
    return myjsonify(out)

###@token_required
###put tokdata as arguemnt in function
@api.route(repo_api_prefix+'/init_repo', methods=['GET'])
def init_repo():
    """
    initialize the repository: HANDLE WITH CARE and have a backup always

    ---
    tags:
      - Utils
    produces:
      - text/html
    responses:
      200:
        description: Successful operation
        examples:
          text/html: 'Repo initialized successfully'
    """
    dbg("++++++++++ entering init_repo")
    #audit(tokdata,request)
    with config.repoengine.connect() as conn:
        pass
    create_repo_db(config.repoengine)
    return 'Repo initialized successfully', 200


###########################
##
## Lookup
##
###########################

@api.route(repo_api_prefix+'/lookup/<id>/data', methods=['GET'])
@token_required
def get_lookup(tokdata,id):
    """
    return then lookup data defined in the lookup repository table with id or alias

    ---
    tags:
      - Repo
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: id
        in: path
        type: string
        required: true
        description: id or alias of the lookup defined in the repository
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering get_lookup")
    dbg_api_call(request)
    dbg("get_lookup: param id is <%s>",str(id))
    audit(tokdata,request)
    out={}
    offset = request.args.get('offset')
    limit = request.args.get('limit')
    order_by = request.args.get('order_by')
    dbg("get_lookup pagination offset=%s limit=%s",offset,limit)
    items,columns,total_count,e=repo_lookup_select(config.repoengine,id,order_by,offset,limit,with_total_count=True,username=tokdata["username"])
    dbg("get_lookup sql_select error %s",str(e))
    if last_stmt_has_errors(e,out):
        try:
            json_out = jsonify(out)
        except:
            err("cannot jsonify "+str(out))
            json_out = ("cannot jsonify "+str(out))[:50]
        return json_out,500
    out["data"]=pre_jsonify_items_transformer(items)
    out["columns"]=columns
    out["total_count"]=total_count
    return myjsonify(out)

###########################
##
## Adhoc
##
###########################
"""
GET /api/repo/adhoc/<id>/data	The data of a adhoc (result of its SQL)
GET /api/repo/adhoc/<id>/data?format=XLSX|CSV	The data of a adhoc (result of its SQL), but as a Excel (XLSX) or CSV file
"""

@api.route(repo_api_prefix+'/adhoc/<id>/data', methods=['GET'])
@token_required
def get_adhoc_data(tokdata,id):
    """
    return then adhoc data defined in the adhoc repository table with id or alias

    ---
    tags:
      - Adhoc
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: id
        in: path
        type: string
        required: true
        description: id or alias of the adhoc defined in the repository
      - name: params
        in: query
        type: string
        description: adhoc parameter
      - name: format
        in: query
        type: string
        description: output format JSON/XLSX/CSV
      - name: offset
        in: query
        type: integer
        description: start with row <offset> (for pagination)
      - name: limit
        in: query
        type: integer
        description: maximum number of rows to return  (for pagination)
      - name: order_by
        in: query
        type: string
        description: order by clause
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering get_adhoc_data")
    dbg_api_call(request)
    dbg("get_adhoc_data: param id is <%s>",str(id))
    prof=get_profile(config.repoengine,tokdata['username'])
    user_id=prof["user_id"]
    out={}
    myparams=None
    fmt="JSON"
    dbg("get_adhoc_data: check request arguments")
    if len(request.args) > 0:
        for key, value in request.args.items():
            dbg("arg: %s val: %s",key,value)
            if key=="format":
                fmt=value
                dbg("adhoc format %s",fmt)
            if key=="params":
                myparams = {}
                slist=value.split(",")
                for s in slist:
                    p=s.split(":")
                    if len(p)>1:
                        myparams[p[0]]=p[1]
                    else:
                        return "adhoc json parameter is invalid, does not contain semicolon",500

    dbg("get_adhoc_data: get request data")
    data_bytes = request.get_data()
    dbg("get_adhoc_data: databytes: %s",data_bytes)
    dataitem = None
    if data_bytes is not None:
        dbg("get_adhoc_data: databytes is not None: %s",data_bytes)
        if len(data_bytes)>0:
            dbg("get_adhoc_data: databytes len > 0: %s",data_bytes)
            data_string = data_bytes.decode('utf-8')
            dbg("get_adhoc_data: datastring: %s",data_string)
            if data_string is not None:
                dataitem = json.loads(data_string)
                dbg("get_adhoc_data: dataitem: %s",str(dataitem))

    offset = request.args.get('offset')
    limit = request.args.get('limit')
    order_by = request.args.get('order_by')
    dbg("get_adhoc_data pagination offset=%s limit=%s",offset,limit)
    dbg("get_adhoc_data pagination order_by=%s",order_by)
    dbg("get_adhoc_data: get adhoc stmt")
    get_rep_adhoc_res = get_repo_adhoc_sql_stmt(config.repoengine,id,user_id)
    if "error" in get_rep_adhoc_res.keys():
        return myjsonify(get_rep_adhoc_res), 500
    adhoc_sql = get_rep_adhoc_res["sql"]
    adhoc_datasrc_id = get_rep_adhoc_res["datasrc_id"]
    adhocid  = get_rep_adhoc_res["adhocid"]
    order_by_def  = get_rep_adhoc_res["order_by_def"]
    adhoc_desc  = get_rep_adhoc_res["adhocdesc"]
    audit(tokdata,request,id=adhocid)
    if adhoc_datasrc_id is None:
        msg="adhoc datasource_id is not set - assuming 1"
        adhoc_datasrc_id = 1
        #log.warning(msg) 
        #return msg, 500
    dbg("get_adhoc_data: parameter substitution")
    # substitute params
    if isinstance(myparams,dict):
        for p,v in myparams.items():
            adhoc_sql=adhoc_sql.replace("$("+p+")",v)
        dbg("get_adhoc_data: adhoc sql after subsitution: %s",adhoc_sql)
    # substitute global environment params
    adhoc_sql=adhoc_sql.replace("$(APP_USER)",tokdata['username'])
    # substitute request data
    if isinstance(dataitem,dict):
        for p,v in dataitem.items():
            adhoc_sql=adhoc_sql.replace("$("+p+")",v)
    dbg("get_adhoc_data: adhoc sql after data subsitution: %s",adhoc_sql)
    if adhoc_sql is None:
        msg="adhoc id/name invalid oder kein sql beim adhoc hinterlegt"
        err(msg)
        return msg, 500
    dbg("get_adhoc_data: get db type")
    adhoc_dbengine = get_db_by_id_or_alias(adhoc_datasrc_id)
    db_typ = get_db_type(adhoc_dbengine)
    dbg("get_adhoc_data: prepare json pagination")
    if fmt=="JSON":
        dbg("get_adhoc_data: fmt JSON")
        adhoc_sql= f"select x.* from ({adhoc_sql}) x"
        adhoc_sql += add_offset_limit(db_typ,offset,limit,order_by)
        dbg("get_adhoc_data JSON pagination: %s",adhoc_sql)
        dbg("get_adhoc_data pagination offset=%s limit=%s",offset,limit)
    else:
        dbg("get_adhoc_data: not fmt JSON/HTML")
        if order_by_def is not None:
            dbg("get_adhoc_data: apply default order by (order by added)")
            adhoc_sql+=" order by "+order_by_def.replace(":"," ")
    #
    # handle formats
    dbg("get_adhoc_data: fmt= %s",fmt)
    if fmt=="JSON":
        # execute adhoc sql
        dbg("get_adhoc_data: execute adhoc sql")
        try:
            items, columns = db_exec(adhoc_dbengine,adhoc_sql)
        except SQLAlchemyError as e_sqlalchemy:
            err("adhoc_sql_errors: %s", str(e_sqlalchemy))
            if last_stmt_has_errors(e_sqlalchemy, out):
                out["error"]+="-get_adhoc_data"
                out["message"]+=" beim Lesen der Adhoc Daten"
            return myjsonify(out), 500
        except Exception as e:
            err("get_adhoc_data exception: %s ",str(e))
            if last_stmt_has_errors(e, out):
                out["error"]+="-get_adhoc_data"
                out["message"]+=" beim Lesen der Adhoc Daten"
            return myjsonify(out), 500
        dbg("get_adhoc_data: fmt JSON")
        if not isinstance(items,list):
            return "adhoc json result error",500
        total_count=len(items)
        out["data"]=pre_jsonify_items_transformer(items)
        out["columns"]=columns
        out["total_count"]=total_count
        return myjsonify(out)
    else:
        dbg("get_adhoc_data: other formats")
        # read data with pandas
        try:
            dbg("adhoc_dbengine %s",str(adhoc_dbengine))
            with adhoc_dbengine.connect() as conn:
                dbg("adhoc_dbengine querying")
                df = pd.read_sql_query(adhoc_sql,conn)
                dbg("adhoc_dbengine query done")
        except SQLAlchemyError as e_sqlalchemy:
            err("adhoc_sql_errors(pd): %s", str(e_sqlalchemy))
            if last_stmt_has_errors(e_sqlalchemy, out):
                out["error"]+="-get_adhoc_data(pd)"
                out["message"]+=" beim Lesen der Adhoc Daten"
            return myjsonify(out), 500
        except Exception as e:
            err("get_adhoc_data exception(pd): %s ",str(e))
            if last_stmt_has_errors(e, out):
                out["error"]+="-get_adhoc_data(pd)"
                out["message"]+=" beim Lesen der Adhoc Daten"
            return myjsonify(out), 500

        #dbg("get_adhoc_data: items=%s",str(items))
        dbg("adhoc_dbengine got pandas dataframe")
        if len(df)==0:
            out["error"]="adhoc-no-rows"
            out["message"]="Die Adhoc Abfrage liefert keine Daten"
            out["detail"]="Die Adhoc Abfrage liefert keine Daten"
            dbg("get_adhoc_data: no rows result")
            return myjsonify(out),500
        else:
            try:
                # Save the DataFrame to an Excel file
                if fmt=="XLSX":
                    dbg("get_adhoc_data: XLSX format")
                    dbg("adhoc excel")
                    tmpfile=os.path.join(tempfile.gettempdir(),'mydata'+datetime.now().strftime("%Y%m%d_%H%M%S")+'.xlsx')
                    datasheet_name="daten"
                    infosheet_name="info"
                    try:
                        output = pd.ExcelWriter(tmpfile,engine="xlsxwriter")
                        output.book.set_properties({"encoding":"utf-8"})
                        fmt_xl.header_style = None
                        #pd.formats.format.header_style = None
                        dbg("get_adhoc_data: df to excel")
                        df.to_excel(output, index=False, sheet_name=datasheet_name)
                        output.close()
                    except Exception as e0:
                        err("get_adhoc_data to_excel exception: %s ",str(e0))
                        out["error"]="get-adhoc-data-toxls"
                        out["message"]="Fehler beim Prozessieren der Adhoc-Daten für den Download (XLSX)"
                        out["detail"]=str(e0)
                        err(traceback.format_exc())
                        log.exception(e0)
                        return myjsonify(out), 500
                    # add sheet with sql
                    book = load_workbook(tmpfile)
                    #autofit columns
                    dbg("get_adhoc_data: add autofit volumns")
                    sheet = book[datasheet_name]
                    sheet_tab = Table(displayName="daten", ref=sheet.dimensions)
                    #default font
                    dbg("get_adhoc_data: default xls font")
                    deffont = Font(name='Arial', size=9, bold=False, italic=False)
                    for row in sheet.iter_rows():
                        for cell in row:
                            cell.font = deffont
                    #header font
                    dbg("get_adhoc_data: header xls font")
                    font = Font(name='Arial', size=9, bold=True, italic=False)
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = (max_length + 2) * 1.2  # Zusätzlicher Puffer und Skalierungsfaktor für die Breite
                        sheet.column_dimensions[column_letter].width = adjusted_width
                        sheet[f'{column_letter}1'].font = font
                    # Iterate over each column and set the filter
                    #sheet.auto_filter.ref = sheet.dimensions
                    sheet.add_table(sheet_tab)
                    #for col_num in range(1, sheet.max_column + 1):
                    #    column_letter = get_column_letter(col_num)
                    #    column_range = f'{column_letter}1:{column_letter}{sheet.max_row}'
                    #    sheet.auto_filter.ref = column_range
                    # Create a new sheet "info"
                    dbg("get_adhoc_data: add info sheet")
                    book.create_sheet(title=infosheet_name)
                    new_sheet = book[infosheet_name]
                    new_sheet['A1'] = "erstellt am:"
                    new_sheet['A2'] = "adhoc:"
                    new_sheet['A3'] = "description:"
                    new_sheet['B1'] = str(datetime.now())
                    new_sheet['B2'] = id
                    new_sheet['B3'] = adhoc_desc
                    # set info field fonts also to Airal 9
                    for f in ['A1','A2','A3','B1','B2','B3']:
                        new_sheet[f].font = deffont
                    #autofit columns
                    for column in new_sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = (max_length + 2) * 1.2  # Zusätzlicher Puffer und Skalierungsfaktor für die Breite
                        new_sheet.column_dimensions[column_letter].width = adjusted_width                    
                    # new sql sheet
                    dbg("get_adhoc_data: add sql sheet")
                    book.create_sheet(title="sql")
                    sql_sheet = book["sql"]
                    sql_sheet.sheet_state = 'hidden'
                    sql_sheet['A1'] = "sql:"
                    sql_sheet['A2'] = adhoc_sql

                    book.save(tmpfile)                    
                    dbg("get_adhoc_data: xlsx saved")
                    # Return the Excel file as a download
                    with open(tmpfile, 'rb') as file:
                        response = Response(
                            file.read(),
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            headers={'Content-Disposition': 'attachment;filename=mydata.xlsx'}
                        )
                        dbg("get_adhoc_data: return response")
                        dbg(response)
                        return response
                elif fmt=="CSV":
                    dbg("adhoc csv")
                    tmpfile='mydata.csv'
                    # Prepare the CSV file
                    try:
                        df.to_csv(tmpfile, index=False)
                    except Exception as e0:
                        err("get_adhoc_data to_csv exception: %s ",str(e0))
                        out["error"]="get-adhoc-data-tocsv"
                        out["message"]="Fehler beim Prozessieren der Adhoc-Daten für den Download (CSV)"
                        out["detail"]=str(e0)
                        err(traceback.format_exc())
                        log.exception(e0)
                        return myjsonify(out), 500
                    # Return the Excel file as a download
                    with open(tmpfile, 'rb') as file:
                        response = Response(
                            file.read(),
                            mimetype='text/csv',
                            headers={'Content-Disposition': 'attachment;filename=mydata.csv'}
                        )
                        dbg(response)
                        return response
                elif fmt=="TXT":
                    dbg("adhoc txt separated with tabs")
                    tmpfile='mydata.csv'
                    # Prepare the CSV file
                    df.to_csv(tmpfile, index=False, sep='\t', quoting=csv.QUOTE_NONE)
                    # Return the Excel file as a download
                    with open(tmpfile, 'rb') as file:
                        response = Response(
                            file.read(),
                            mimetype='text/csv',
                            headers={'Content-Disposition': 'attachment;filename=mydata.csv'}
                        )
                        dbg(response)
                        return response
                else: 
                    out["error"]="adhoc-invalid-format"
                    out["message"]="Das Format des Adhocs muss XLSX/CSV/TXT/JSON sein"
                    out["detail"]=None
                    return myjsonify(out), 500
            except Exception as e:
                err("get_adhoc_data exception: %s ",str(e))
                out["error"]="get-adhoc-data-fai"
                out["message"]="Fehler beim Prozessieren der Adhoc-Daten für den Download"
                out["detail"]=str(e)
                return myjsonify(out), 500
    out["error"]="get_adhoc_data-should-not-occur"
    out["message"] = "adhoc error that should not happen"
    return myjsonify(out), 500

users=dict()

def load_repo_users():
    """
    load all users defined in the repository into a global dictionary "user"
    i.e. caching for performance reasons
    """
    dbg("++++++++++ entering load_repo_users")
    global users
    out={}
    plainbi_users,columns,cnt,e=sql_select(config.repoengine,'plainbi_user')
    if last_stmt_has_errors(e,out):
        err('error in select users %s', str(e))
        return False
    users = {u["username"]: { "password_hash": u["password_hash"], "email" : u["email"], "rolename" : "Admin" if u["role_id"]==1 else "User" } for u in plainbi_users}

def get_user_by_email(emailname):
    """
    find a user by his email
    i.e. caching for performance reasons
    """
    global users
    dbg("++++++++++ entering get_user_by_email for %s",emailname)
    if not isinstance(emailname,str):
        log.warning("calling get_user_by_email with no string")
        return None
    for unam,u in users.items():
      if "email" in u.keys():
          m = u["email"]
          if isinstance(m,str):
              if m.lower() == emailname.lower():
                  return unam
    return None

def load_repo_users():
    """
    load all users defined in the repository into a global dictionary "user"
    i.e. caching for performance reasons
    """
    dbg("++++++++++ entering load_repo_users")
    global users
    out={}
    plainbi_users,columns,cnt,e=sql_select(config.repoengine,'plainbi_user')
    if last_stmt_has_errors(e,out):
        err('error in select users %s', str(e))
        return False
    users = {u["username"]: { "password_hash": u["password_hash"], "email" : u["email"], "rolename" : "Admin" if u["role_id"]==1 else "User" } for u in plainbi_users}


def authenticate_local(username,password):
    """
    authenticate a local (repository) user
    """
    dbg("++++++++++ entering authenticate_local")
    dbg_api_call(request)
    global users
    load_repo_users()
    if not username or not password:
        err('error invalid cred')
        return False

    p=config.bcrypt.generate_password_hash(password)
    pwd_hashed=p.decode()
    dbg("login: hashed input pwd is %s",pwd_hashed)
    if username in users.keys():
        if config.bcrypt.check_password_hash(users[username]["password_hash"], password):
            dbg("login: pwd ok")
            return True
    else:
        dbg("login: user %s is unknown in repo",username)
    dbg("++++++++++ leaving login")
    return False


def authenticate_ldap(login_username,password=None):
    """
    authenticate a user via LDAP Active Directory
    login_username can be ntaccount (cn) or email (mail)
    if no password is then just find the user in the LDAP 
    """
    dbg("++++++++++ entering authenticate_ldap")
    dbg_api_call(request)
    global users
    mail=None
    full_name=None
    load_repo_users()
    authenticated=False
    bindpwd=os.environ.get("LDAP_BIND_USER_PASSWORD")
    bindpwd=bindpwd.strip()
    username=login_username
    dbg("login username from ldap=%s",username)
    s = ldap3.Server(host=os.environ.get("LDAP_HOST"), port=int(os.environ.get("LDAP_PORT")), use_ssl=False, get_info=ldap3.ALL)
    conn_bind = ldap3.Connection(s, user=os.environ.get("LDAP_BIND_USER_DN"), password=bindpwd, auto_bind='NONE', version=3, authentication='SIMPLE')
    if not conn_bind.bind():
        err('error in bind %s', str(conn_bind.result))
        err('check environent variables LDAP_HOST=%s LDAP_PORT=%s LDAP_BIND_USER_DN=%s', os.environ.get("LDAP_HOST"),os.environ.get("LDAP_PORT"),os.environ.get("LDAP_BIND_USER_DN"))
        if "LDAP_BIND_USER_PASSWORD" not in list(dict(os.environ).keys()):
            err("environment variable LDAP_BIND_USER_PASSWORD is missing")
        dbg("++++++++++ entering authenticate_ldap with status %s",authenticated)
        return authenticated,username
    if "LDAP_BASE_DN" not in list(dict(os.environ).keys()):
        err("environment variable LDAP_BASE_DN is missing")
        return authenticated,username
    if os.environ.get("LDAP_SEARCH_EXPR") is not None:
        search_expr=os.environ.get("LDAP_SEARCH_EXPR")
        search_expr=search_expr.replace("{username}",username)
    else:
        if "@" in username:
            # login by email adress
            search_expr=f'(mail={username})'
        else:
            search_expr=f'(&(cn={username}))'
    dbg("LDAP Search Expression is %s",search_expr)
    conn_bind.search(os.environ.get("LDAP_BASE_DN"), search_expr, attributes=['*'])
    for entry in conn_bind.entries:
        dbg("ldap entry=%s",entry.entry_dn)
        # substitute username by returned cn
        #username=entry.cn.value
        try:
            username=entry.sAMAccountName.value.lower()
        except Exception as e:
            dbg("authenticate_ldap:%s",str(e))
            return authenticated, username
        dbg("username (sAMAccountName) from ldap=%s",username)
        if password is not None:
            # validate the login with thepassword
            conn_auth = ldap3.Connection(s, user=entry.entry_dn, password=password, auto_bind='NONE', version=3, authentication='SIMPLE')
            if not conn_auth.bind():
                log.warning("error in bind ldap entry=%s",entry.entry_dn)
                authenticated=False
            else:
                authenticated=True
        # add user to repository
        if username not in users.keys():
            log.warning("new user %s from ldap registered",username)
            mail = entry.mail.value if 'mail' in entry else None
            full_name = entry.displayName.value if 'displayName' in entry else None
            db_adduser(config.repoengine,username,pwd=None,is_admin=False,email=mail,fullname=full_name)
            dbg("refresh profile cache")
            config.profile_cache={}
        break # user was found in ldap, no need to search more
    dbg("++++++++++ entering authenticate_ldap with status %s",authenticated)
    return authenticated,username


@api.route('/login', methods=['POST'])
@api.route('/api/login', methods=['POST'])
def login():
    """
    User login, authenticate a user - login procedure
    try LDAP first if it is configured (environment variables)
    otherwise of if no success try local authentication
    summary: login to plainbi backend (Active Directory LDAP or internal user management)
    If the login is successful one can enter the returned access token into the dialog of the Swagger Authorize button. Afterwards you can try out the protected endpoints
    ---
    tags:
      - Authentication
    description: Login endpoint for user authentication
    consumes:
      - "application/json"
    parameters:
      -  name: body
         in: body
         required: true
         schema:
            required:
              - username
              - password
            properties:
              username:
                type: string
                description: User's username (in LDAP AD email is also possible)
                example: "admin"
              password:
                type: string
                description: User's password
    responses:
      200:
        description: Successful login
        examples:
          application/json: 
            "access_token": "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VybmFtZSI6ImFkbWluIn0.w08k-KbwtT8DphvaFEn0Ruwf6Px0pGoSh1-E9UakpyE"
            "message": "Login erfolgreich"
            "role": "Admin"
      401:
        description: Unauthorized
        examples:
          application/json: 
            "detail": "invalid-credentials in local auth"
            "error": "invalid-credentials"
            "message": "Benutzername oder Passwort ist falsch"
    """
    out={}
    dbg("++++++++++ entering login")
    dbg_api_call(request)
    dbg("login")
    data_bytes = request.get_data()
    referer = request.headers.get('Referer')
    dbg("login referer is %s",str(referer))
    username = None # init
    #dbg("databytes: %s",data_bytes)
    data_string = data_bytes.decode('utf-8')
    #dbg("datastring: %s",data_string)
    item = json.loads(data_string.strip("'"))
    #print("login items ",str(item))

    login_username = item['username'].lower()
    dbg("login: username=%s",login_username)
    password = item['password']
    if len(login_username)==0:
        out["message"]='Username muss angegeben werden'
        out["error"]="empty-credentials"
        out["detail"]="invalid-credentials no username"
        return myjsonify(out), 401
    if len(password)==0:
        out["message"]='Passwort darf nicht leer sein'
        out["error"]="empty-credentials"
        out["detail"]="invalid-credentials no password"
        return myjsonify(out), 401

    #dbg("login: password=%s",password)
    #audit(item['username'],request)
    audit(item['username'],request)

    used_ldap=False
    used_local=False
    authenticated = False
    if "LDAP_HOST" in list(dict(os.environ).keys()):  # if LDAP is defined in environment
        used_ldap=True
        authenticated,username = authenticate_ldap(login_username,password)
        dbg("login authenticated by ldap = %s",authenticated)
        if not authenticated:
            dbg("try locally authenticated")
            username = login_username  # use original name in login mask for local auth
            authenticated = authenticate_local(username,password)
            dbg("login authenticated local = %s",authenticated)
    else:
        username = login_username 
        dbg("ldap authentication skipped because no LDAP_HOST environment variable")
        used_local=True
        authenticated = authenticate_local(username,password)
        dbg("login authenticated local = %s",authenticated)
    if authenticated:
        dbg("login authenticated")
        token = jwt.encode({'username': username}, config.SECRET_KEY, algorithm='HS256')
        if username not in users.keys():
            dbg('refresh users array')
            load_repo_users()
        if len(request.args) > 0:
            for key, value in request.args.items():
                dbg("arg: %s val: %s",key,value)
                if key=="tokenonly":  # this helps for testing
                    return token
        else:
            return myjsonify({'access_token': token, "message":"Login erfolgreich", 'role': users[username]["rolename"]}), 200
    else:
        out["message"]='Benutzername oder Passwort ist falsch'
        out["error"]="invalid-credentials"
        if used_ldap and used_local:
            out["detail"]="invalid-credentials in ldap and local auth"
        elif used_ldap:
            out["detail"]="invalid-credentials in ldap auth"
        elif used_local:
            out["detail"]="invalid-credentials in local auth"
        else:
            out["detail"]="invalid-credentials without ldap and local"
    return myjsonify(out), 401



@api.route('/login_sso', methods=['POST'])
@api.route('/api/login_sso', methods=['POST'])
def login_sso():
    """
    User login with sso, authenticate a user
    ---
    tags:
      - Authentication
    description: Login endpoint for user authentication with SSO
    responses:
      200:
        description: Successful login
        examples:
          application/json: 
            "access_token": "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VybmFtZSI6ImFkbWluIn0.w08k-KbwtT8DphvaFEn0Ruwf6Px0pGoSh1-E9UakpyE"
            "message": "Login erfolgreich"
            "role": "Admin"
      401:
        description: Unauthorized
        examples:
          application/json: 
            "detail": "invalid-credentials in login_sso"
            "error": "invalid-credentials"
            "message": "Benutzername oder Passwort ist falsch"
    """
    out={}
    dbg("++++++++++ entering login_sso")
    dbg_api_call(request)
    dbg("login_sso")
    data_bytes = request.get_data()
    referer = request.headers.get('Referer')
    dbg("login referer is %s",str(referer))
    #audit(item['username'],request)
    used_ldap=False
    used_local=False
    authenticated = False
    data_string = data_bytes.decode('utf-8')
    dbg("login_sso datastring: %s",data_string,dbglevel=3)
    item = json.loads(data_string.strip("'"))
    dbg("login_sso item: %s",str(item))
    dbg("login_sso state: %s",item.get("state"))
    dbg("now validate data we've got from Microsoft")
    
    dbg("calling auth2 token")
    token_url = f"https://login.microsoftonline.com/{config.PLAINBI_SSO_TENANTID}/oauth2/v2.0/token" 
    dbg("token url is: %s",token_url)
    dbg("config.PLAINBI_SSO_REDIRECT_PATH is: %s",config.PLAINBI_SSO_REDIRECT_PATH)
    #payload = { 'grant_type': 'authorization_code', 'code' : item["code"], 'redirect_uri' : 'http://localhost:5000/getSSOToken',
    payload = { 'grant_type': 'authorization_code', 'code' : item["code"], 'redirect_uri' : config.PLAINBI_SSO_REDIRECT_PATH, 
                'client_id' : config.PLAINBI_SSO_APPLICATION_ID,  'scope' : "User.Read",    'client_secret' :  config.PLAINBI_SSO_CLIENT_SECRET
    }
    response = requests.post(token_url, data=payload)
    dbg("response= "+str(response.text))
    tokens = response.json()
    dbg("tokens= "+str(tokens))
    ms_id_token = tokens['id_token']
    i_ms_key=0
    for ms_key in config.ms_keys:
        try:
            i_ms_key+=1
            #log.info("validate id token try %d ",i_ms_key)
            decoded_id_token = jwt.decode(ms_id_token, key=jwt.algorithms.RSAAlgorithm.from_jwk(ms_key), algorithms=['RS256'], audience = config.PLAINBI_SSO_APPLICATION_ID, issuer=f"https://login.microsoftonline.com/{config.PLAINBI_SSO_TENANTID}/v2.0")
            #dbg("decoded_id_token is %s",str(decoded_id_token))
            useremail = decoded_id_token["preferred_username"]
            username = get_user_by_email(useremail)
            if username is None:
                # try to add user from ldap to plainbi
                warn(f"User {useremail} is not yet in the user table. Try to find it in LDAP and create the user")
                authenticate_ldap(useremail,password=None)
                dbg('refresh users array')
                load_repo_users()
                username = get_user_by_email(useremail)
                if username is None:
                    warn("user %s is not in know users",str(username))
            if username is not None: 
                dbg("username is %s",str(username))
                log.info("Valid Id Token: User unique name: %s",username)
                authenticated = True
            else:
                authenticated = False
                warn("No Valid Id Token : %s",username)
            break
        #except (jwt.InvalidTokenError,jwt.DecodeError):
        except Exception as e_validate:
            warn("validing id token failed with %s",str(e_validate))
            continue
    else:
      err("calling auth2 token: ID Token validation failed")

    if authenticated:
        dbg("login authenticated by SSO for %s",username)
        # create a new token for plainbi web
        token = jwt.encode({'username': username}, config.SECRET_KEY, algorithm='HS256')
        if username not in users.keys():
            dbg('refresh users array')
            load_repo_users()
        if len(request.args) > 0:
            for key, value in request.args.items():
                dbg("arg: %s val: %s",key,value)
                if key=="tokenonly":  # this helps for testing
                    return token
        else:
            dbg("++++++++++ leaving login_sso authenticated ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            return myjsonify({'access_token': token, "message":"Login erfolgreich", 'role': users[username]["rolename"]}), 200
    else:
        dbg("login NOT authenticated")
        out["message"]='SSO Login war nicht erfolgreich'
        out["error"]="sso invalid-credentials"
        if used_ldap and used_local:
            out["detail"]="invalid-credentials in ldap and local auth"
        elif used_ldap:
            out["detail"]="invalid-credentials in ldap auth"
        elif used_local:
            out["detail"]="invalid-credentials in local auth"
        else:
            out["detail"]="invalid-credentials without ldap and local"
    return myjsonify(out), 401


@api.route('/passwd', methods=['POST'])
@api.route('/api/passwd', methods=['POST'])
@token_required
def passwd(tokdata):
    """
    change a local users password 

    ---
    tags:
      - Authentication
    security:
    - APIKeyHeader: ['Authorization']
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    out={}
    dbg("passwd")
    dbg_api_call(request)
    data_bytes = request.get_data()
    dbg("databytes: %s",data_bytes)
    data_string = data_bytes.decode('utf-8')
    dbg("datastring: %s",data_string)
    item = json.loads(data_string.strip("'"))
    dbg("passwd items ",str(item),dbglevel=3)
    prof=get_profile(config.repoengine,tokdata['username'])
    
    plainbi_users,columns,cnt,e=sql_select(config.repoengine,'plainbi_user')
    if last_stmt_has_errors(e,out):
        return myjsonify({'error': 'Invalid User collecting'}), 500
    users = {u["username"]: u["password_hash"] for u in plainbi_users}
    dbg(str(users),dbglevel=3)

    password = item['password']
    dbg("login: password=%s",password)
    p=config.bcrypt.generate_password_hash(password)
    pwd_hashed=p.decode()
    dbg(pwd_hashed,dbglevel=3)
    
    if prof["role"] == "Admin":
        username = item['username']
        dbg("passwd: username=%s",username)
    else:
        username=prof["username"]
        oldpassword = item['old_password']
        dbg("login: password=%s",oldpassword)
        if username in users.keys():
            if config.bcrypt.check_password_hash(users[username], oldpassword):
                dbg("old pwd ok")
                out["error"]="old-password-does-not-match"
                out["message"]="Altes Passwort ist falsch"
                return myjsonify(out)
    out=db_passwd(config.repoengine,username,p)
    dbg("++++++++++ leaving passwd with %s",out)
    return myjsonify(out)


@api.route('/hash_passwd/<pwd>', methods=['GET'])
@api.route('/api/hash_passwd/<pwd>', methods=['GET'])
def hash_passwd(pwd):
    """
    just show the hashed password ... mainly for testing reasons

    ---
    tags:
      - Utils
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    out={}
    out["pwd"]=pwd
    #p=config.bcrypt.generate_password_hash(pwd.encode('utf-8'))
    #pwd_hashed=p.decode()
    p=config.bcrypt.generate_password_hash(pwd)
    pwd_hashed=p.decode()
    out["hashed"]=pwd_hashed
    dbg("hashed pwd: "+pwd_hashed,dbglevel=3)
    return myjsonify(out)

@api.route('/cache', methods=['GET'])
@api.route('/api/cache', methods=['GET'])
@token_required
def cache(tokdata):
    """
    cache handling of metadata, profile
    url params
      on .... enable caching
      off ... disable caching
      clear ... clear caching
      status ... show current cache handling setting

    returns simple string and status 200

    ---
    tags:
      - Misc
    security:
    - APIKeyHeader: ['Authorization']
    produces:
      - text/plain
    responses:
      200:
        description: Successful operation
        examples:
          text/plain: "cache is enabled/disabled"
    """
    dbg_api_call(request)
    config.metadataraw_cache={}
    config.profile_cache={}
    dbg("clear_cache: get_metadata_raw: cache created")
    dbg("clear_cache: get_profile: cache created")
    if len(request.args) > 0:
        for key, value in request.args.items():
            dbg("arg: %s val: %s",key,value)
            if key=="on":
                config.use_cache=True
                dbg("caching enabled")
                config.metadataraw_cache = {}
                config.profile_cache = {}
                return 'cacheing endabled', 200
            if key=="off":
                config.use_cache=False
                dbg("caching disabled")
                return 'cacheing disabled', 200
            if key=="clear":
                config.metadataraw_cache={}
                config.profile_cache={}
                dbg("clear_cache: get_metadata_raw: cache created")
                dbg("clear_cache: get_profile: cache created")
                return 'caches cleared', 200
            if key=="status":
                if config.use_cache:
                    return 'cache is enabled', 200
                else:
                    return 'cache is disabled', 200

    return 'caches cleared', 200

@api.route('/clear_cache', methods=['GET'])
@api.route('/api/clear_cache', methods=['GET'])
@token_required
def clear_cache(tokdata):
    """
    clear caches (metadata and profile cache)
    returns simple string and status 200

    ---
    tags:
      - Misc
    produces:
      - text/plain
    security:
    - APIKeyHeader: ['Authorization']
    responses:
      200:
        description: Successful operation
        examples:
          text/plain: 'caches cleared' 
    """
    dbg_api_call(request)
    config.metadataraw_cache={}
    config.profile_cache={}
    dbg("clear_cache: get_metadata_raw: cache cleared")
    dbg("clear_cache: get_profile: cache cleared")
    return 'caches cleared', 200

@api.route('/protected', methods=['GET'])
@api.route('/api/protected', methods=['GET'])
@token_required
def protected(tokdata):
    """
    show the own username

    ---
    tags:
      - Misc
    description: "show your own username"
    security:
    - APIKeyHeader: ['Authorization']
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    dbg("current user=%s",tokdata['username'])
    u=tokdata['username']
    return myjsonify({'message': f'Hello, {u}! You are authenticated.'}), 200

@api.route('/profile', methods=['GET'])
@api.route('/api/profile', methods=['GET'])
@token_required
def profile(tokdata):
    """
    return json of the profile of the current user

    ---
    tags:
      - Misc
    security:
    - APIKeyHeader: ['Authorization']
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
    """
    audit(tokdata,request)
    out=get_profile(config.repoengine,tokdata['username'])
    return myjsonify(out)


@api.route('/logout', methods=['GET'])
@api.route('/api/logout', methods=['GET'])
def logout(tokdata):
    """
    logout

    ---
    tags:
      - Authentication
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: logged out
    """
    dbg("logout")
    audit(tokdata,request)
    return myjsonify({'message': 'logged out'})

# dsdb export
@api.route(repo_api_prefix+'/application/<appid>/dsdb', methods=['GET'])
#@token_required
#def download_app_dsdb(tokdata,appid):
def download_app_dsdb(appid):
    """
    download a dsdb file for the application object in the repository
    can/should be used for deployments

    ---
    tags:
      - Misc
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: appid
        in: path
        type: string
        required: true
        description: id or alias of the application defined in the repository
      - name: filenam
        in: query
        type: string
        description: output filename
    responses:
      200:
        description: Successful operation
        examples:
          text/plain: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering download_app_dsdb")
    dbg_api_call(request)
    dbg("download_app_dsdb: app_id is <%s>",str(appid))
    out=get_item_raw(config.repoengine, "plainbi_application", str(appid))
    if "data" in out.keys():
        print("app=",str(out))
        rec=out["data"][0]
        s='{\n  objectList:\n  [\n    {\n      dsdbFormat: 1\n      deploymentType: always\n      current:\n      {\n        version: 1.0\n'
        s+='        statements: [\n          DELETE FROM plainbi_application WHERE id IN ('+str(rec["id"])+');\n        ]\n        data:\n        [\n'
        s+='          {\n            target: plainbi_application\n            columns: [ "id", "name", "alias", "datasource_id", "spec_json"]\n            rows: [\n'
        s+='              [\n                '+str(rec["id"])+', "'+rec["name"]+'", "'+rec["alias"]+'", '+str(rec["datasource_id"])+'\n'
        s+="                '''\n"
        s+= rec["spec_json"]
        s+="\n                '''\n"
        s+="              ]\n            ]\n          }\n        ]\n      }\n    }\n  ]\n}\n"
        # Return the data as a download
        response = Response(
            s,
            mimetype='text/plain',
            headers={'Content-Disposition': 'attachment; filename=mydata.dsdb'}
        )
        dbg(response)
        return response
    else:
        return "error getting application or application does not exist", 500


@api.route(repo_api_prefix+'/lookup/<lkpid>/dsdb', methods=['GET'])
#@token_required
#def download_lkp_dsdb(tokdata,lkpid):
def download_lkp_dsdb(lkpid):
    """
    download a dsdb file for the lookup object in the repository
    can/should be used for deployments

    ---
    tags:
      - Misc
    security:
    - APIKeyHeader: ['Authorization']
    parameters:
      - name: lkpid
        in: path
        type: string
        required: true
        description: id or alias of the application defined in the repository
      - name: filenam
        in: query
        type: string
        description: output filename
    responses:
      200:
        description: Successful operation
        examples:
          text/plain: 
            message: Data processed successfully
    """
    dbg("++++++++++ entering download_lkp_dsdb")
    dbg_api_call(request)
    dbg("download_lkp_dsdb: app_id is <%s>",str(lkpid))
    out=get_item_raw(config.repoengine, "plainbi_lookup", lkpid)
    if "data" in out.keys():
        rec=out["data"][0]
        print("lkp=",str(out))
        s='{\n  objectList:\n  [\n    {\n      dsdbFormat: 1\n      deploymentType: always\n      current:\n      {\n        version: 1.0\n'
        s+='        statements: [\n          DELETE FROM plainbi_lookup WHERE id IN ('+str(rec["id"])+');\n        ]\n        data:\n        [\n'
        s+='          {\n            target: plainbi_lookup\n            columns: [ "id", "name", "alias", "datasource_id", "sql_query"]\n            rows: [\n'
        s+='              [\n                '+str(rec["id"])+', "'+rec["name"]+'", "'+rec["alias"]+'", '+str(rec["datasource_id"])+'\n'
        s+="                '''\n"
        s+=rec["sql_query"]
        s+="\n                '''\n"
        s+="              ]\n            ]\n          }\n        ]\n      }\n    }\n  ]\n}\n"
        # Return the data as a download
        response = Response(
            s,
            mimetype='text/plain',
            headers={'Content-Disposition': 'attachment; filename=mydata.dsdb'}
        )
        dbg(response)
        return response
        dbg("++++++++++ entering download_lkp_dsdb")
        dbg_api_call(request)
        dbg("download_lkp_dsdb: lookup_id is <%s>",str(appid))
        s = "Hallo\nhugo"
        # Return the data as a download
        response = Response(
            s,
            mimetype='text/plain',
            headers={'Content-Disposition': 'attachment; filename=mydatalkp.dsdb'}
        )
        dbg(response)
        return response
    else:
        return "error getting lookup or lookup does not exist", 500



###########################
##
## Static
##
###########################


@api.route('/api/static/<id>', methods=['GET'])
@api.route('/static/<id>', methods=['GET'])
def getstatic(id):
    """
    gets a static base64 thing from the repo by id or alias without login
    useful for logo etc.
    base table is plainbi_static_file

    ---
    tags:
      - Utils
    produces:
      - text/plain
    parameters:
      - name: id
        in: path
        type: string
        required: true
        description: id or alias of the static object defined in the repository
    responses:
      200:
        description: Successful operation
        examples:
          text/plain: 'base64 string of object/image etc.' 
      404:
        description: static object not found
        examples:
          text/plain: 'no data found' 
    """
    dbg_api_call(request)
    if is_id(id):
        sql_params={ "id" : id}
        sql="select * from plainbi_static_file where id=:id"
    else:
        sql_params={ "alias" : id}
        sql="select * from plainbi_static_file where alias=:alias"
    dbg("getstatic: sql is <%s>",sql)
    s,s_columns = db_exec(config.repoengine, sql , sql_params)
    #dbg("static resource = %s",str(s))
    if len(s)>0:
        for r in s:
            b64 = r["content_base64"]
            response = make_response(base64.b64decode(b64))
            response.headers.set('Content-Type', r["mimetype"])
            #response.headers.set('Content-Disposition', 'attachment', filename='%s.jpg' % pid)
            return response
    else:
        return "no data found",404

@api.route('/api/settings.js', methods=['GET'])
def getsettingsjs():
    """
    base table is plainbi_setting

    ---
    tags:
      - Utils
    produces:
      - text/javascript
    responses:
      200:
        description: Successful operation
        examples:
          text/javascript: "var APP_TITLE = ...."
    """
    global app
    dbg("++++++++++ entering getsettingsjs")
    dbg_api_call(request)
    
    out={}
    dbg("getsettings from db")
    items,columns,total_count,e=sql_select(config.repoengine,"plainbi_settings",with_total_count=True)
    if isinstance(e,str) and e=="ok":
        dbg("getsettings sql_select ok")
    else:
        dbg("getsettings sql_select error %s",str(e))
    if last_stmt_has_errors(e,out):
        try:
            json_out2 = jsonify(out)
        except Exception as ej2:
            err("getsettings.js: jsonify Error 2: %s",str(ej2))
        return json_out2,500

    def get_setting_from_list(items,nam):
        for i in items:
            if i["setting_name"]==nam:
                if i["setting_value"] is None:
                    return ""
                else:
                    return i["setting_value"]
        return ""

    dbg("construct javascript")
    dbg("settings are %s",str(items))
    s=  "// header and footer\n"
    s=s+f"var APP_TITLE = '"+get_setting_from_list(items,'app_title')+"';\n"
    s=s+f"var HEADER_TITLE = '"+get_setting_from_list(items,'header_title')+"';\n"
    s=s+f"var FOOTER = '"+get_setting_from_list(items,'footer')+"';\n"
    s=s+"\n"
    s=s+"// environment banner\n"
    s=s+f"var ENVIRONMENT_BANNER_TEXT = '"+get_setting_from_list(items,'environment_banner_text')+"'; // e.g. DEV, TEST - leave empty for PROD, as you mostly don't need a banner there\n"
    s=s+"\n"
    s=s+"// theme\n"
    s=s+f"var THEME_COLOR_PRIMARY = '"+get_setting_from_list(items,'color_primary')+"';\n"
    s=s+f"var THEME_COLOR_SUCCESS = '"+get_setting_from_list(items,'color_success')+"';\n"
    s=s+f"var THEME_COLOR_ERROR = '"+get_setting_from_list(items,'color_error')+"';\n"
    s=s+f"var THEME_COLOR_INFO = '"+get_setting_from_list(items,'color_info')+"';\n"
    s=s+f"var THEME_FONT_SIZE = "+get_setting_from_list(items,'font_size')+";\n"
    s=s+f"var CONTACT_EMAIL = '"+get_setting_from_list(items,'contact_email')+"';\n"
    if config.PLAINBI_SSO_APPLICATION_ID is not None:
        dbg("get SSO signin Link")
        #config.PLAINBI_SSO_REDIRECT_PATH
        config.PLAINBI_SSO_SCOPE = ["User.Read"]
        dbg(f"config.PLAINBI_SSO_SCOPE = {config.PLAINBI_SSO_SCOPE}")
        dbg(f"config.PLAINBI_SSO_REDIRECT_PATH = {config.PLAINBI_SSO_REDIRECT_PATH}")
        try:
            ssoapp = msal.ConfidentialClientApplication(client_id=config.PLAINBI_SSO_APPLICATION_ID, authority=config.PLAINBI_SSO_AUTHORITY, client_credential=config.PLAINBI_SSO_CLIENT_SECRET)
            dbg("ssoapp initialized")
            ssourl = ssoapp.get_authorization_request_url(config.PLAINBI_SSO_SCOPE, redirect_uri = config.PLAINBI_SSO_REDIRECT_PATH, state="hugo" )
            dbg(f"sso auth url is: %s",ssourl)
            uri = ssourl
            dbg("auth uri is %s",uri)
            parsed_uri = urlparse(uri)
            query_params = parse_qs(parsed_uri.query)
            vstate=query_params.get('state', [None])[0]
            config.SSO_CODE_CHALLENGE = query_params.get('code_challenge', [None])[0]
            dbg("SSO_CODE_CHALLENGE %s",config.SSO_CODE_CHALLENGE)
            s=s+f"var SSO_SIGNIN_LINK = '"+uri+"';\n"
        except Exception as e_ssoapp:
            err("SSO msal app error: "+str(e_ssoapp))
            config.PLAINBI_SSO_APPLICATION_ID = None
            config.with_sso = False
            log.warning("SSO disabled due to error in msal create app")

    response = make_response(s)
    response.headers.set('Content-Type', "text/javascript; charset=utf-8")
    return response

@api.route('/api/settings', methods=['GET'])
def getsettings():
    """
    get all settings
    base table is plainbi_setting
    ---
    tags:
      - Utils
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            data: Data
            columns: Columns
            total_count: int
      500:
        description: getting setting failed
    """
    out={}
    dbg("++++++++++ entering getsettings")
    dbg_api_call(request)
    items,columns,total_count,e=sql_select(config.repoengine,"plainbi_settings",with_total_count=True)
    if isinstance(e,str) and e=="ok":
        dbg("getsettings sql_select ok")
    else:
        dbg("getsettings sql_select error %s",str(e))
    if last_stmt_has_errors(e,out):
        try:
            json_out2 = jsonify(out)
        except Exception as ej2:
            err("getsettings: jsonify Error 2: %s",str(ej2))
            log.exception(ej2)
        return json_out2,500
    out["data"]=pre_jsonify_items_transformer(items)
    out["columns"]=columns
    out["total_count"]=total_count
    dbg("leaving getsettings and return json result")
    dbg("out=%s",str(out))
    return myjsonify(out)

@api.route('/api/setting/<name>', methods=['GET'])
def getsetting(name):
    """
    get a specific setting value by name
    base table is plainbi_settinggs

    ---
    tags:
      - Utils
    parameters:
      - name: name
        in: path
        type: string
        required: true
        description: name of the setting in the repository
    responses:
      200:
        description: Successful operation
        examples:
          application/json: 
            message: Data processed successfully
      404:
        description: Setting not found
    """
    dbg("++++++++++ entering getsetting")
    sql_params={ "name" : name}
    dbg_api_call(request)
    sql="select * from plainbi_settings where setting_name=:name"
    dbg("getsetting: sql is <%s>",sql)
    s,s_columns = db_exec(config.repoengine, sql , sql_params)
    dbg("setting %s = %s",name,str(s))
    out={}
    if len(s)>0:
        for r in s:
            out["setting_name"] = r["setting_name"]
            out["setting_value"] = r["setting_value"]
            return myjsonify(out)
    else:
        return "no data found", 404

#p_verbose=args.verbose, p_logfile=args.logfile, p_configfile=args.config, p_repository=args.repository, p_database=args.database, p_port=args.port 
def create_app(p_verbose=None, p_logfile=None, p_repository=None, p_database=None, p_port=None):
    """
    create app is the standard Flask application definition

    it is called either from 
      - the standalone plainbi_backend.py 
      - or from the uwsgi script
      - unittest scripts (the parameters p_repository and p_database are important here)
    that's why the get_config handling is necessary
    """
    dbg("++++++++++ entering create_app")
    global app


    log.info("creating flask app")
    app = Flask(__name__)
    app.config["SESSION_PERMANENT"] = True
    if with_swagger:
        log.info("swagger enabled")
        swagger = Swagger(app, template={
            "info" : {
                "title" : "plainbi Backend Flask API",
                "description": "Swagger for plainbi https://github.com/markuskolp/plainbi",
                "version" : config.version
            },
            'securityDefinitions': {
                'APIKeyHeader': {
                        'type': 'apiKey',
                        'name': 'Authorization',
                        'in': 'header'
                }
            }
        }, 
        )

    app.json_encoder = CustomJSONEncoder ## wegen jsonify datetimes
    app.register_blueprint(api)
    
    app.config.from_object(cfg)

    if p_repository:
        app.config["PLAINBI_REPOSITORY"] = p_repository

    dbg(f"app.config.SESSION_TYPE = {app.config['SESSION_TYPE']}")
    Session(app)
   
    # get the configuration
    #get_config(repository=p_repository,database=p_database,verbose=3)
    
    # connect to the repository
    #config.repoengine = db_connect(config.repository)
    config.repoengine = db_connect(app.config["PLAINBI_REPOSITORY"])
    if not db_connect_test(config.repoengine):
        err("cannot connect to repository. Check repository database connection description 'PLAINBI_REPOSITORY' in config file or environment")
        sys.exit(0)

    # get datasources from repository
    log.info("load datasources from plainbi_datasource")        
    load_datasources_from_repo()

    if not config.database:
        try:
           config.database = config.datasources["1"]
        except Exception as e:
            log.warning("config datasource %s",str(e))
            log.exception(e)

    # if there is a database database now connect to it
    if config.database:
        config.dbengine = db_connect(config.database)
        if not db_connect_test(config.dbengine):
            err("cannot connect to database. Check database connection description 'PLAINBI_DATABASE' in config file or environment")
            sys.exit(0)
        log.info(f"The default database connection description is {config.database}")
    
    #from yourapplication.views.admin import admin
    #from yourapplication.views.frontend import frontend
    #app.register_blueprint(admin)
    #app.register_blueprint(frontend)

    # handle Java Web Tokens
    app.config['JWT_SECRET_KEY'] = config.SECRET_KEY
    app.secret_key = config.SECRET_KEY
    config.bcrypt = Bcrypt(app)

    if config.PLAINBI_SSO_APPLICATION_ID is not None:
        log.info("prepare SSO Login")
        #config.PLAINBI_SSO_REDIRECT_PATH
        dbg(f"config.PLAINBI_SSO_AUTHORITY = {config.PLAINBI_SSO_AUTHORITY}")
        dbg(f"config.PLAINBI_SSO_APPLICATION_ID = {config.PLAINBI_SSO_APPLICATION_ID}")
        #dbg(f"config.PLAINBI_SSO_CLIENT_SECRET = {config.PLAINBI_SSO_CLIENT_SECRET}")
        dbg(f"config.PLAINBI_SSO_TENANTID = {config.PLAINBI_SSO_TENANTID}")
        log.info("SSO auth initialized")

        dbg("SSO: get microsoft keys")
        try:
            ms_keys_url = f"https://login.microsoftonline.com/{config.PLAINBI_SSO_TENANTID}/discovery/v2.0/keys"
            config.ms_keys = requests.get(ms_keys_url).json()["keys"]
            #dbg("ms_keys is: %s",str(config.ms_keys))
        except Exception as e_get_ms_keys:
            err("get microsoft keys for SSO: %s",str(e_get_ms_keys))
            err("SSO disabled coz cannot get microsoft keys for tenant %s",config.PLAINBI_SSO_TENANTID)
            config.with_sso=False
            config.ms_keys=[]


    # begin: multi process uwsgi database connection pool handling
    # https://stackoverflow.com/questions/59248806/how-to-correctly-setup-flask-uwsgi-sqlalchemy-to-avoid-database-connection-i
    def _dispose_db_pool():
        with app.app_context():
            config.repoengine.engine.dispose()

    try:
        from uwsgidecorators import postfork
        postfork(_dispose_db_pool)
        log.info(f"uwsgi postfork enabled for repository connection")
    except ImportError:
        # Implement fallback when running outside of uwsgi...
        log.warning(f"uwsgi postfork NOT enabled for repository connection (but maybe because just standalone version)")
    # end: multi process uwsgi database connection pool handling

    #jwt = JWTManager(app)
    return app

